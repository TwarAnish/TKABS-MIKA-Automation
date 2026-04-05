# from django.core.management.base import BaseCommand, CommandError
# from django.conf import settings
# from psr.models import PSRProject, PSRSnapshot, PSRSubDepartment, ProjectCostCategory, ForecastAdjustment, MaterialForecastAdjustment
# from openpyxl import Workbook
# from openpyxl.styles import Font
# from openpyxl.utils import get_column_letter
# from decimal import Decimal
# from datetime import datetime
# import os

# class Command(BaseCommand):
#     help = 'Generate Forecast Excel report for a PSR project (3 sheets: Labor Hours, Labor Costs, Material Costs)'

#     def add_arguments(self, parser):
#         parser.add_argument('--project_id', type=int, help='Project ID')
#         parser.add_argument('--co_no', type=str, help='Project CO number')
#         parser.add_argument('--date', type=str, help='Specific snapshot date in YYYY-MM-DD format (optional; defaults to latest snapshot)')
#         parser.add_argument('--output', type=str, default=None, help='Output file path (default: media/forecast_reports/<co_no>_<date>.xlsx)')

#     def handle(self, *args, **options):
#         if options['project_id']:
#             try:
#                 project = PSRProject.objects.get(pk=options['project_id'])
#             except PSRProject.DoesNotExist:
#                 raise CommandError('Project with given ID does not exist')
#         elif options['co_no']:
#             try:
#                 project = PSRProject.objects.get(co_no=options['co_no'])
#             except PSRProject.DoesNotExist:
#                 raise CommandError('Project with given CO number does not exist')
#         else:
#             raise CommandError('Provide either --project_id or --co_no')

#         snapshots_qs = project.psr_snapshots.all()

#         if options['date']:
#             try:
#                 target_date = datetime.strptime(options['date'], '%Y-%m-%d').date()
#                 try:
#                     current_snapshot = snapshots_qs.get(snapshot_date=target_date)
#                 except PSRSnapshot.DoesNotExist:
#                     raise CommandError(f'No snapshot found for date {options["date"]}')
#             except ValueError:
#                 raise CommandError('Invalid date format. Use YYYY-MM-DD')
#         else:
#             if not snapshots_qs.exists():
#                 raise CommandError('No snapshots found for this project')
#             current_snapshot = snapshots_qs.order_by('-snapshot_date').first()

#         # Determine previous snapshot (the one before the current one)
#         previous_snapshot = (
#             snapshots_qs
#             .filter(snapshot_date__lt=current_snapshot.snapshot_date)
#             .order_by('-snapshot_date')
#             .first()
#         )

#         latest = current_snapshot  # for consistency with existing code
#         previous = previous_snapshot

#         # Hardcoded order as specified
#         labor_codes = [
#             "PM", "POM", "PEM", "KMA/KHP", "DET", "DOK", "2D", "SIM", "QC", "LAY",
#             "PEC", "KEL", "KES", "IBS", "IBSS", "IBK", "PAM", "MMA/MHP", "A+I(M)", "INS", "A+I(E)"
#         ]
#         material_codes = [
#             "KTFT", "KTES", "KTEP", "KTMA", "KTHP", "ZKE", "F+V", "SOKO", "RK",
#             "ASSEMBLE SERVICES", "DESIGN SERVICES", "FACTORY EQUIPMENTS CONSUMABLES",
#             "STATIONARY", "QC (QUALITY CHECKING SERVICES)"
#         ]

#         wb = Workbook()

#         # ==================== Helper to find leaf by code ====================
#         def find_leaf_by_code(nested_dict, target_code):
#             if isinstance(nested_dict, dict):
#                 if target_code in nested_dict and isinstance(nested_dict[target_code], dict) and "id" in nested_dict[target_code]:
#                     return nested_dict[target_code]
#                 for value in nested_dict.values():
#                     found = find_leaf_by_code(value, target_code)
#                     if found:
#                         return found
#             return None

#         # ==================== Sheet creation helper ====================
#         def create_sheet(sheet_name, is_labor, is_hours):
#             ws = wb.create_sheet(title=sheet_name)
#             header_col_a = "Subdepartment" if is_labor else "Cost Category"

#             ws['A1'] = header_col_a
#             ws['B1'] = "Lines"
#             ws['C1'] = "Line Value"
#             ws['D1'] = "Current Month Forecast Value"
#             ws['E1'] = "Last Month Forecast Value"

#             for cell in ws[1]:
#                 cell.font = Font(bold=True)

#             # Number formats
#             if not is_hours:  # Costs - add ₹
#                 for col in ['C', 'D', 'E']:
#                     ws.column_dimensions[col].number_format = '"₹"#,##0.00'
#             else:
#                 for col in ['C', 'D', 'E']:
#                     ws.column_dimensions[col].number_format = '#,##0.00'

#             return ws

#         # ==================== Labor Hours Sheet ====================
#         ws_hours = wb.active
#         ws_hours.title = "Labor Hours"
#         ws_hours = create_sheet("Labor Hours", is_labor=True, is_hours=True)

#         timesheet_hours_latest = latest.data.get("TIMESHEET", {}).get("HOURS", {})
#         timesheet_hours_prev = previous.data.get("TIMESHEET", {}).get("HOURS", {}) if previous else {}

#         row = 2
#         for code in labor_codes:
#             leaf_latest = find_leaf_by_code(timesheet_hours_latest, code)
#             leaf_prev = find_leaf_by_code(timesheet_hours_prev, code) if previous else None

#             current_base = float(leaf_latest.get("rest", 0.0)) if leaf_latest else 0.0
#             prev_forecast = float(leaf_prev.get("rest", 0.0)) if leaf_prev else 0.0

#             subdept = None
#             if leaf_latest and "id" in leaf_latest:
#                 try:
#                     subdept = PSRSubDepartment.objects.get(pk=leaf_latest["id"])
#                 except PSRSubDepartment.DoesNotExist:
#                     pass

#             lines = []
#             current_forecast = current_base
#             if subdept and subdept.forecast_override:
#                 adjustment = subdept.forecast_adjustments.order_by('-adjustment_date').first()
#                 if adjustment:
#                     lines = list(adjustment.lines.all())
#                     if lines:
#                         current_forecast = sum(float(line.hours) for line in lines)

#             if lines:
#                 first = True
#                 for line in lines:
#                     if first:
#                         ws_hours[f'A{row}'] = code
#                         ws_hours[f'A{row}'].font = Font(bold=True)
#                         ws_hours[f'D{row}'] = round(current_forecast, 2)
#                         ws_hours[f'E{row}'] = round(prev_forecast, 2)
#                         first = False
#                     ws_hours[f'B{row}'] = line.description or ""
#                     ws_hours[f'C{row}'] = round(float(line.hours), 2)
#                     row += 1
#             else:
#                 ws_hours[f'A{row}'] = code
#                 ws_hours[f'A{row}'].font = Font(bold=True)
#                 ws_hours[f'B{row}'] = "-"
#                 ws_hours[f'C{row}'] = "-"
#                 ws_hours[f'D{row}'] = round(current_forecast, 2)
#                 ws_hours[f'E{row}'] = round(prev_forecast, 2)
#                 row += 1

#             row += 1  # blank row after group

#         # ==================== Labor Costs Sheet ====================
#         ws_costs = create_sheet("Labor Costs", is_labor=True, is_hours=False)

#         timesheet_cost_latest = latest.data.get("TIMESHEET", {}).get("COST", {})
#         timesheet_cost_prev = previous.data.get("TIMESHEET", {}).get("COST", {}) if previous else {}

#         row = 2
#         for code in labor_codes:
#             leaf_latest = find_leaf_by_code(timesheet_cost_latest, code)
#             leaf_prev = find_leaf_by_code(timesheet_cost_prev, code) if previous else None

#             current_base = float(leaf_latest.get("rest", 0.0)) if leaf_latest else 0.0
#             prev_forecast = float(leaf_prev.get("rest", 0.0)) if leaf_prev else 0.0

#             subdept = None
#             rate_inr = Decimal('0.0')
#             if leaf_latest and "id" in leaf_latest:
#                 try:
#                     subdept = PSRSubDepartment.objects.get(pk=leaf_latest["id"])
#                     rate_inr = subdept.department.hourly_rate * project.exchange_rate
#                 except PSRSubDepartment.DoesNotExist:
#                     pass

#             lines = []
#             current_forecast = current_base
#             if subdept and subdept.forecast_override:
#                 adjustment = subdept.forecast_adjustments.order_by('-adjustment_date').first()
#                 if adjustment:
#                     lines = list(adjustment.lines.all())
#                     if lines:
#                         current_forecast = sum(float(line.hours) * float(rate_inr) for line in lines)

#             if lines:
#                 first = True
#                 for line in lines:
#                     line_value = float(line.hours) * float(rate_inr)
#                     if first:
#                         ws_costs[f'A{row}'] = code
#                         ws_costs[f'A{row}'].font = Font(bold=True)
#                         ws_costs[f'D{row}'] = round(current_forecast, 2)
#                         ws_costs[f'E{row}'] = round(prev_forecast, 2)
#                         first = False
#                     ws_costs[f'B{row}'] = line.description or ""
#                     ws_costs[f'C{row}'] = round(line_value, 2)
#                     row += 1
#             else:
#                 ws_costs[f'A{row}'] = code
#                 ws_costs[f'A{row}'].font = Font(bold=True)
#                 ws_costs[f'B{row}'] = "-"
#                 ws_costs[f'C{row}'] = "-"
#                 ws_costs[f'D{row}'] = round(current_forecast, 2)
#                 ws_costs[f'E{row}'] = round(prev_forecast, 2)
#                 row += 1

#             row += 1  # blank row

#         # ==================== Material Costs Sheet ====================
#         ws_material = create_sheet("Material Costs", is_labor=False, is_hours=False)

#         cost_to_go_latest = latest.data.get("COST TO GO", {}).get("COST", {})
#         cost_to_go_prev = previous.data.get("COST TO GO", {}).get("COST", {}) if previous else {}

#         row = 2
#         for code in material_codes:
#             leaf_latest = find_leaf_by_code(cost_to_go_latest, code)
#             leaf_prev = find_leaf_by_code(cost_to_go_prev, code) if previous else None

#             current_base = float(leaf_latest.get("rest", 0.0)) if leaf_latest else 0.0
#             prev_forecast = float(leaf_prev.get("rest", 0.0)) if leaf_prev else 0.0

#             pcc = None
#             if leaf_latest and "id" in leaf_latest:
#                 try:
#                     pcc = ProjectCostCategory.objects.get(pk=leaf_latest["id"])
#                 except ProjectCostCategory.DoesNotExist:
#                     pass

#             lines = []
#             current_forecast = current_base
#             if pcc and pcc.forecast_override:
#                 adjustment = pcc.forecast_adjustments.order_by('-adjustment_date').first()
#                 if adjustment:
#                     lines = list(adjustment.lines.all())
#                     if lines:
#                         current_forecast = sum(float(line.amount) for line in lines)

#             if lines:
#                 first = True
#                 for line in lines:
#                     if first:
#                         ws_material[f'A{row}'] = code
#                         ws_material[f'A{row}'].font = Font(bold=True)
#                         ws_material[f'D{row}'] = round(current_forecast, 2)
#                         ws_material[f'E{row}'] = round(prev_forecast, 2)
#                         first = False
#                     ws_material[f'B{row}'] = line.description or ""
#                     ws_material[f'C{row}'] = round(float(line.amount), 2)
#                     row += 1
#             else:
#                 ws_material[f'A{row}'] = code
#                 ws_material[f'A{row}'].font = Font(bold=True)
#                 ws_material[f'B{row}'] = "-"
#                 ws_material[f'C{row}'] = "-"
#                 ws_material[f'D{row}'] = round(current_forecast, 2)
#                 ws_material[f'E{row}'] = round(prev_forecast, 2)
#                 row += 1

#             row += 1  # blank row

#         # Remove default sheet
#         wb.remove(wb["Labor Hours"])

#         # Auto-adjust column widths
#         for ws in wb.worksheets:
#             for col in ws.columns:
#                 max_length = 0
#                 column = col[0].column_letter
#                 for cell in col:
#                     if cell.value:
#                         max_length = max(max_length, len(str(cell.value)))
#                 adjusted_width = min(max_length + 2, 50)
#                 ws.column_dimensions[column].width = adjusted_width

#         # Save file
#         if options['output']:
#             file_path = options['output']
#         else:
#             media_root = settings.MEDIA_ROOT
#             reports_dir = os.path.join(media_root, 'forecast_reports')
#             os.makedirs(reports_dir, exist_ok=True)
#             snapshot_date_str = latest.snapshot_date.strftime('%Y%m%d')
#             file_name = f"Forecast_Report_{project.co_no}_{snapshot_date_str}.xlsx"
#             file_path = os.path.join(reports_dir, file_name)

#         wb.save(file_path)
#         self.stdout.write(self.style.SUCCESS(f'Forecast report generated for snapshot {latest.snapshot_date}: {file_path}'))




from django.core.management.base import BaseCommand, CommandError
from django.conf import settings
from psr.models import PSRProject, PSRSnapshot, PSRSubDepartment, ProjectCostCategory, ForecastAdjustment, MaterialForecastAdjustment
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from decimal import Decimal
from datetime import datetime
import os

class Command(BaseCommand):
    help = 'Generate Forecast Excel report for a PSR project (3 sheets: Labor Hours, Labor Costs, Material Costs)'

    def add_arguments(self, parser):
        parser.add_argument('--project_id', type=int, help='Project ID')
        parser.add_argument('--co_no', type=str, help='Project CO number')
        parser.add_argument('--date', type=str, help='Specific snapshot date in YYYY-MM-DD format (optional; defaults to latest snapshot)')
        parser.add_argument('--output', type=str, default=None, help='Output file path (default: media/forecast_reports/<co_no>_<date>.xlsx)')

    def handle(self, *args, **options):
        if options['project_id']:
            try:
                project = PSRProject.objects.get(pk=options['project_id'])
            except PSRProject.DoesNotExist:
                raise CommandError('Project with given ID does not exist')
        elif options['co_no']:
            try:
                project = PSRProject.objects.get(co_no=options['co_no'])
            except PSRProject.DoesNotExist:
                raise CommandError('Project with given CO number does not exist')
        else:
            raise CommandError('Provide either --project_id or --co_no')

        snapshots_qs = project.psr_snapshots.all()

        if options['date']:
            try:
                target_date = datetime.strptime(options['date'], '%Y-%m-%d').date()
                try:
                    current_snapshot = snapshots_qs.get(snapshot_date=target_date)
                except PSRSnapshot.DoesNotExist:
                    raise CommandError(f'No snapshot found for date {options["date"]}')
            except ValueError:
                raise CommandError('Invalid date format. Use YYYY-MM-DD')
        else:
            if not snapshots_qs.exists():
                raise CommandError('No snapshots found for this project')
            current_snapshot = snapshots_qs.order_by('-snapshot_date').first()

        # Determine previous snapshot (the one before the current one)
        previous_snapshot = (
            snapshots_qs
            .filter(snapshot_date__lt=current_snapshot.snapshot_date)
            .order_by('-snapshot_date')
            .first()
        )

        latest = current_snapshot  # for consistency with existing code
        previous = previous_snapshot

        # Hardcoded order as specified
        labor_codes = [
            "PM", "POM", "PEM", "KMA/KHP", "DET", "DOK", "2D", "SIM", "QC", "LAY",
            "PEC", "KEL", "KES", "IBS", "IBSS", "IBK", "PAM", "MMA/MHP", "A+I(M)", "INS", "A+I(E)"
        ]
        material_codes = [
            "KTFT", "KTES", "KTEP", "KTMA", "KTHP", "ZKE", "F+V", "SOKO", "RK",
            "ASSEMBLE SERVICES", "DESIGN SERVICES", "FACTORY EQUIPMENTS CONSUMABLES",
            "STATIONARY", "QC (QUALITY CHECKING SERVICES)"
        ]

        wb = Workbook()

        # ==================== Helper to find leaf by code ====================
        def find_leaf_by_code(nested_dict, target_code):
            if isinstance(nested_dict, dict):
                if target_code in nested_dict and isinstance(nested_dict[target_code], dict) and "id" in nested_dict[target_code]:
                    return nested_dict[target_code]
                for value in nested_dict.values():
                    found = find_leaf_by_code(value, target_code)
                    if found:
                        return found
            return None

        # ==================== Sheet creation helper ====================
        def create_sheet(sheet_name, is_labor, is_hours):
            ws = wb.create_sheet(title=sheet_name)
            header_col_a = "Subdepartment" if is_labor else "Cost Category"

            ws['A1'] = header_col_a
            ws['B1'] = "Lines"
            ws['C1'] = "Line Value"
            ws['D1'] = "Current Forecast Value"
            ws['E1'] = "Previous Forecast Value"

            for cell in ws[1]:
                cell.font = Font(bold=True)

            # Number formats
            if not is_hours:  # Costs - add ₹
                for col in ['C', 'D', 'E']:
                    ws.column_dimensions[col].number_format = '"₹"#,##0.00'
            else:
                for col in ['C', 'D', 'E']:
                    ws.column_dimensions[col].number_format = '#,##0.00'

            return ws

        # ==================== Labor Hours Sheet ====================
        ws_hours = wb.active
        ws_hours.title = "Labor Hours"
        ws_hours = create_sheet("Labor Hours", is_labor=True, is_hours=True)

        timesheet_hours_latest = latest.data.get("TIMESHEET", {}).get("HOURS", {})
        timesheet_hours_prev = previous.data.get("TIMESHEET", {}).get("HOURS", {}) if previous else {}

        row = 2
        for code in labor_codes:
            leaf_latest = find_leaf_by_code(timesheet_hours_latest, code)
            leaf_prev = find_leaf_by_code(timesheet_hours_prev, code) if previous else None

            current_base = float(leaf_latest.get("rest", 0.0)) if leaf_latest else 0.0
            prev_forecast = float(leaf_prev.get("rest", 0.0)) if leaf_prev else 0.0

            subdept = None
            if leaf_latest and "id" in leaf_latest:
                try:
                    subdept = PSRSubDepartment.objects.get(pk=leaf_latest["id"])
                except PSRSubDepartment.DoesNotExist:
                    pass

            lines = []
            current_forecast = current_base
            if subdept and subdept.forecast_override:
                adjustment = subdept.forecast_adjustments.order_by('-adjustment_date').first()
                if adjustment:
                    lines = list(adjustment.lines.all())
                    if lines:
                        current_forecast = sum(float(line.hours) for line in lines)

            if lines:
                first = True
                for line in lines:
                    if first:
                        ws_hours[f'A{row}'] = code
                        ws_hours[f'A{row}'].font = Font(bold=True)
                        ws_hours[f'D{row}'] = round(current_forecast, 2)
                        ws_hours[f'E{row}'] = round(prev_forecast, 2)
                        first = False
                    ws_hours[f'B{row}'] = line.description or ""
                    ws_hours[f'C{row}'] = round(float(line.hours), 2)
                    row += 1
            else:
                ws_hours[f'A{row}'] = code
                ws_hours[f'A{row}'].font = Font(bold=True)
                ws_hours[f'B{row}'] = "-"
                ws_hours[f'C{row}'] = "-"
                ws_hours[f'D{row}'] = round(current_forecast, 2)
                ws_hours[f'E{row}'] = round(prev_forecast, 2)
                row += 1

            row += 1  # blank row after group

        # ==================== Labor Costs Sheet ====================
        ws_costs = create_sheet("Labor Costs", is_labor=True, is_hours=False)

        timesheet_cost_latest = latest.data.get("TIMESHEET", {}).get("COST", {})
        timesheet_cost_prev = previous.data.get("TIMESHEET", {}).get("COST", {}) if previous else {}

        row = 2
        for code in labor_codes:
            leaf_latest = find_leaf_by_code(timesheet_cost_latest, code)
            leaf_prev = find_leaf_by_code(timesheet_cost_prev, code) if previous else None

            current_base = float(leaf_latest.get("rest", 0.0)) if leaf_latest else 0.0
            prev_forecast = float(leaf_prev.get("rest", 0.0)) if leaf_prev else 0.0

            subdept = None
            rate_inr = Decimal('0.0')
            if leaf_latest and "id" in leaf_latest:
                try:
                    subdept = PSRSubDepartment.objects.get(pk=leaf_latest["id"])
                    rate_inr = subdept.department.hourly_rate * project.exchange_rate
                except PSRSubDepartment.DoesNotExist:
                    pass

            lines = []
            current_forecast = current_base
            if subdept and subdept.forecast_override:
                adjustment = subdept.forecast_adjustments.order_by('-adjustment_date').first()
                if adjustment:
                    lines = list(adjustment.lines.all())
                    if lines:
                        current_forecast = sum(float(line.hours) * float(rate_inr) for line in lines)

            if lines:
                first = True
                for line in lines:
                    line_value = float(line.hours) * float(rate_inr)
                    if first:
                        ws_costs[f'A{row}'] = code
                        ws_costs[f'A{row}'].font = Font(bold=True)
                        ws_costs[f'D{row}'] = round(current_forecast, 2)
                        ws_costs[f'E{row}'] = round(prev_forecast, 2)
                        first = False
                    ws_costs[f'B{row}'] = line.description or ""
                    ws_costs[f'C{row}'] = round(line_value, 2)
                    row += 1
            else:
                ws_costs[f'A{row}'] = code
                ws_costs[f'A{row}'].font = Font(bold=True)
                ws_costs[f'B{row}'] = "-"
                ws_costs[f'C{row}'] = "-"
                ws_costs[f'D{row}'] = round(current_forecast, 2)
                ws_costs[f'E{row}'] = round(prev_forecast, 2)
                row += 1

            row += 1  # blank row

        # ==================== Material Costs Sheet ====================
        ws_material = create_sheet("Material Costs", is_labor=False, is_hours=False)

        cost_to_go_latest = latest.data.get("COST TO GO", {}).get("COST", {})
        cost_to_go_prev = previous.data.get("COST TO GO", {}).get("COST", {}) if previous else {}

        row = 2
        for code in material_codes:
            leaf_latest = find_leaf_by_code(cost_to_go_latest, code)
            leaf_prev = find_leaf_by_code(cost_to_go_prev, code) if previous else None

            current_base = float(leaf_latest.get("rest", 0.0)) if leaf_latest else 0.0
            prev_forecast = float(leaf_prev.get("rest", 0.0)) if leaf_prev else 0.0

            pcc = None
            if leaf_latest and "id" in leaf_latest:
                try:
                    pcc = ProjectCostCategory.objects.get(pk=leaf_latest["id"])
                except ProjectCostCategory.DoesNotExist:
                    pass

            lines = []
            current_forecast = current_base
            if pcc and pcc.forecast_override:
                adjustment = pcc.forecast_adjustments.order_by('-adjustment_date').first()
                if adjustment:
                    lines = list(adjustment.lines.all())
                    if lines:
                        current_forecast = sum(float(line.amount) for line in lines)

            if lines:
                first = True
                for line in lines:
                    if first:
                        ws_material[f'A{row}'] = code
                        ws_material[f'A{row}'].font = Font(bold=True)
                        ws_material[f'D{row}'] = round(current_forecast, 2)
                        ws_material[f'E{row}'] = round(prev_forecast, 2)
                        first = False
                    ws_material[f'B{row}'] = line.description or ""
                    ws_material[f'C{row}'] = round(float(line.amount), 2)
                    row += 1
            else:
                ws_material[f'A{row}'] = code
                ws_material[f'A{row}'].font = Font(bold=True)
                ws_material[f'B{row}'] = "-"
                ws_material[f'C{row}'] = "-"
                ws_material[f'D{row}'] = round(current_forecast, 2)
                ws_material[f'E{row}'] = round(prev_forecast, 2)
                row += 1

            row += 1  # blank row

        # Remove default sheet
        wb.remove(wb["Labor Hours"])

        # Auto-adjust column widths
        for ws in wb.worksheets:
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

        # Save file
        if options['output']:
            file_path = options['output']
        else:
            media_root = settings.MEDIA_ROOT
            reports_dir = os.path.join(media_root, 'forecast_reports')
            os.makedirs(reports_dir, exist_ok=True)
            snapshot_date_str = latest.snapshot_date.strftime('%Y%m%d')
            file_name = f"Forecast_Report_{project.co_no}_{snapshot_date_str}.xlsx"
            file_path = os.path.join(reports_dir, file_name)

        wb.save(file_path)
        self.stdout.write(self.style.SUCCESS(f'Forecast report generated for snapshot {latest.snapshot_date}: {file_path}'))