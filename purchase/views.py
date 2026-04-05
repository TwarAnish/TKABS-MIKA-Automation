from openpyxl.drawing.image import Image
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from rest_framework.views import APIView
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from rest_framework.decorators import api_view,permission_classes, authentication_classes
from rest_framework.response import Response
from rest_framework import status, permissions
from rest_framework import status
from rest_framework.permissions import AllowAny
from decimal import Decimal, InvalidOperation
from .models import *
from .serializers import *
from accounts.models import *
from django.utils import timezone
from django.db import transaction
from datetime import datetime
from purchase.models import Project as PurchaseProject
from capacity_planning.models import Project as CapacityProject
from django.core.serializers.json import DjangoJSONEncoder
from datetime import date
from django.contrib.auth import get_user_model
from django.core.exceptions import ObjectDoesNotExist, MultipleObjectsReturned
from rest_framework_simplejwt.tokens import RefreshToken
from django.db.models import Sum, Count
from django.db.models.functions import TruncMonth
import pandas as pd
import numpy as np
import logging
logger = logging.getLogger(__name__)
from rest_framework.parsers import MultiPartParser, FormParser
from django.shortcuts import get_object_or_404
from rest_framework.views import APIView
from rest_framework.permissions import AllowAny
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Count, Sum, Q
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from django.db.models import Sum
from django.db.models.functions import ExtractYear, ExtractMonth
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.csrf import csrf_exempt
from collections import defaultdict
#project api's
import json
from rest_framework import status, parsers
def get_tokens_for_user(user):
  refresh = RefreshToken.for_user(user)
  return {
      'refresh': str(refresh),
      'access': str(refresh.access_token),
  }
 
HYDRATION_MAP = {
    'projectboard': HydrateProjectBoardSerializer,
    'commercialrequirement': HydrateCommercialSerializer,
    'supplier': HydrateSupplierSerializer, # Use previous logic for supplier
    'lineitem': HydrateLineItemSerializer,
    'termassignment': HydrateTermAssignmentSerializer,
}

class DashboardDataView(APIView):
    authentication_classes = []
    permission_classes = [AllowAny]

    def get(self, request):
        try:
            data = Project.objects.aggregate(
                active_projects=Count("id", filter=Q(status=Status.ACTIVE)),
                total_actual_expense=Sum("actual_expense"),
            )

            data = {
                "active_projects": data["active_projects"] or 0,
                "total_actual_expense": data["total_actual_expense"] or 0,
                "total approved PB froms": 0,
                "total non-approved PB froms": 0,
            }

            return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"Error Fetching Data"}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
 
class ProjectListCreateAPIView(APIView):
    parser_classes = [MultiPartParser, FormParser]

    def get(self, request):
        user = request.user
        queryset = Project.objects.all().select_related('department', 'project_manager')

        # Check if the user is assigned as a manager to ANY project
        is_a_manager = Project.objects.filter(project_manager=user).exists()

        # Logic: If they are a manager (and not a superuser), restrict them.
        # Otherwise (Admins, Staff, or users who aren't managers), show all.
        if is_a_manager and not user.is_superuser:
            projects = queryset.filter(project_manager=user)
        else:
            projects = queryset

        serializer = ProjectCreateSerializer(projects, many=True)
        return Response(serializer.data)

    def post(self, request):
        # 1. Prepare Data
        serializer = ProjectCreateSerializer(data=request.data)
        
        # 2. Validate
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # 3. Save with Atomic Transaction
        try:
            with transaction.atomic():
                project = serializer.save(created_by=request.user if request.user.is_authenticated else None)

                # Handle Excel if present
                attachment = request.FILES.get("attachment")
                if attachment:
                    # upload_session = ProcurementUpload.objects.create(
                    #     project=project, 
                    #     excel=attachment, 
                    #     is_approved=True 
                    # )
                    self.handle_excel_data(attachment, project)

            return Response(ProjectCreateSerializer(project).data, status=status.HTTP_201_CREATED)
        
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    
    def handle_excel_data(self, file, project):
        try:
            xl = pd.ExcelFile(file)
            target_sheet = "Procurement Plan "
            
            # 1. Flexible Sheet Selection
            if target_sheet not in xl.sheet_names:
                sheets_lower = {s.lower().strip(): s for s in xl.sheet_names}
                if "Procurement Plan " in sheets_lower:
                    target_sheet = sheets_lower["procurement_plan "]
                else:
                    raise Exception(f"Sheet '{target_sheet}' not found. Available: {xl.sheet_names}")

            # 2. Parse with the correct header row
            # header=2 refers to the 3rd row in Excel (0-indexed). 
            # Adjust to 0, 1, or 2 based on where the 'Item' column starts.
            df = xl.parse(sheet_name=target_sheet, header=3)

            # 3. Clean 'Unnamed' columns and empty rows
            df = df.dropna(how="all")
            df.columns = df.columns.astype(str).str.strip()
            df = df.loc[:, ~df.columns.str.contains("^Unnamed", na=False)]

            # 4. Normalize columns carefully
            # This turns "Qty Required" into "qty_required"
            # And "Delivery Date required @ tkAB (Planned)" into "delivery_date_required_@_tkab_planned"
            df.columns = (
                df.columns
                .str.lower()
                .str.replace(r'[\(\)\.]', '', regex=True) # Remove (), .
                .str.replace(r'\s+', '_', regex=True)     # Replace spaces with underscores
                .str.strip('_')
            )

            # 5. Column Validation
            # Match these exactly to the result of your normalization above
            date_col = "delivery_date_required_@_tkab_planned"
            required = ["item", "qty_required", "budget", date_col]
            
            missing = [col for col in required if col not in df.columns]
            if missing:
                raise Exception(f"Missing columns: {missing}. Available normalized: {list(df.columns)}")

            # 6. Data Cleaning
            df = df.replace([np.nan, np.inf, -np.inf], 0)

            # 7. Create Records
            excel_records = []
            for _, row in df.iterrows():
                item_name = str(row.get("item", "")).strip()
                if not item_name or item_name == "0":
                    continue

                # Process Date
                raw_date = row.get(date_col)
                formatted_date = None
                if isinstance(raw_date, pd.Timestamp):
                    formatted_date = raw_date.date()
                elif isinstance(raw_date, str) and raw_date.strip():
                    formatted_date = parse_date(raw_date)

                excel_records.append(
                    ExcelFile(
                        project=project,
                        item=item_name,
                        total=row["qty_required"], # Mapped from Excel
                        item_budget=row["budget"], # Mapped from Excel
                        delivery_date=formatted_date,
                        is_active=True
                    )
                )

            if excel_records:
                ExcelFile.objects.bulk_create(excel_records)

        except Exception as e:
            raise Exception(f"Excel Error: {str(e)}")

class ProjectDetailAPIView(APIView):
    # Add authentication/permission classes as needed
    
    def get_object(self, pk):
        return get_object_or_404(Project, pk=pk)

    def get(self, request, pk):
        project = self.get_object(pk)
        serializer = ProjectCreateSerializer(project)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, pk):
        project = self.get_object(pk)
        serializer = ProjectCreateSerializer(project, data=request.data, partial=True)
        
        if serializer.is_valid():
            try:
                # Capture values for comparison
                new_budget = serializer.validated_data.get("total_budget")
                old_budget = project.total_budget
                user = request.user if request.user.is_authenticated else None

                with transaction.atomic():
                    # 1. Update project fields from validated data
                    # serializer.save() works here, but since you want ALL logic in view:
                    for attr, value in serializer.validated_data.items():
                        setattr(project, attr, value)
                    project.save()

                    # 2. Handle Budget History logic
                    if new_budget is not None and new_budget != old_budget:
                        ProjectBudget.objects.create(
                            project=project,
                            budget=new_budget,
                            created_by=user
                        )

                # Use ListSerializer for the response to include calculated fields (deviation, etc.)
                response_serializer = ProjectCreateSerializer(project)
                return Response(response_serializer.data, status=status.HTTP_200_OK)

            except Exception as e:
                return Response(
                    {"error": "Update failed", "detail": str(e)}, 
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class ProjectNameListAPIView(APIView):
    """Returns a list of capacity project names and codes."""
    def get(self, request):
        projects = MainProject.objects.all().values("id", "project_name", "co_no")
        
        data = [
            {
                "id": p["id"],
                "name": p["project_name"],
                "project_code": p["co_no"],
            }
            for p in projects
        ]
        
        return Response({
            "count": len(data),
            "results": data
        }, status=status.HTTP_200_OK)


class ProjectBudgetPatchView(APIView):
    def patch(self, request, pk):
        # This gets the record regardless of project type
        project_instance = get_object_or_404(Project, pk=pk)
        
        # We use partial=True so we only update the budget
        serializer = ProjectBudgetUpdateSerializer(
            project_instance, 
            data=request.data, 
            partial=True
        )
        
        if serializer.is_valid():
            serializer.save()
            return Response({
                "status": "success",
                "message": "Budget updated",
                "new_budget": serializer.data['total_budget']
            }, status=status.HTTP_200_OK)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Value, F, Q
from django.db.models.functions import Concat
from collections import defaultdict

from rest_framework import generics

class OrganizationalUsersAPIView(APIView):
    """
    Returns categorized users in a single response based on roles and departments.
    """
    def get(self, request):
        # 1. Single query fetching only needed fields
        users = CustomUser.objects.filter(is_active=True).annotate(
            full_name=Concat(F('first_name'), Value(' '), F('last_name'))
        ).values("id", "username", "full_name", "role", "department__name")

        # Initialize structure to ensure keys exist even if lists are empty
        response_data = {
            "managing_directors": [],
            "project_managers": [],
            "hods": [],
            "buyers": [],
            "employees": [], # General category for non-purchase employees
        }

        for user in users:
            user_info = {
                "id": user["id"],
                "username": user["username"],
                "name": user["full_name"].strip() or user["username"],
                "dept": user["department__name"]
            }
            
            role = (user["role"] or "").lower()
            dept_name = (user["department__name"] or "").lower()

            # 1. MDs, PMs, and HODs
            if role == "md":
                response_data["managing_directors"].append(user_info)
            elif role == "project_manager":
                response_data["project_managers"].append(user_info)
            elif role == "hod":
                response_data["hods"].append(user_info)
            
            # 2. Employees (Split by Department)
            elif role == "employee":
                if dept_name == "purchase":
                    response_data["buyers"].append(user_info)
                else:
                    response_data["employees"].append(user_info)

        return Response(response_data, status=status.HTTP_200_OK)
 
 

class CashInflowListCreateAPIView(APIView):
    # permission_classes = [IsAuthenticated] # Uncomment when ready

    def get(self, request):
        """
        Fetches cash inflows with related project details.
        """
        inflows = CashInflow.objects.select_related(
            "project", 
            "project__project"
        ).order_by("-date")

        results = []
        for inflow in inflows:
            # inflow.project is the Procurement Project
            # inflow.project.project is the Core Capacity Project
            core_project = inflow.project.project 

            results.append({
                "id": inflow.id,
                "procurement_project_id": inflow.project_id,
                "core_project_id": core_project.id if core_project else None,
                "project_code": core_project.co_no if core_project else None,
                "project_name": core_project.project_name if core_project else None,
                "amount": float(inflow.amount),
                "date": inflow.date,
                "status": inflow.status,
                "created_at": inflow.created_at,
            })

        return Response({
            "count": len(results),
            "results": results
        }, status=status.HTTP_200_OK)

    def post(self, request):
        """
        Handles creation of cash inflows with status validation logic.
        """
        # Role Check Logic
        # if request.user.role != "Project Manager":
        #     return Response({"error": "Only project managers can add cash inflow"}, status=status.HTTP_403_FORBIDDEN)

        serializer = CashInflowCreateSerializer(data=request.data)
        
        if serializer.is_valid():
            # Business Logic: Only ACTIVE projects allowed
            project_instance = serializer.validated_data.get('project')
            
            # Note: Ensure 'Status.ACTIVE' is imported or defined
            if project_instance.status != Status.ACTIVE:
                return Response(
                    {"project": "Cash inflow can be added only to ACTIVE projects."},
                    status=status.HTTP_400_BAD_REQUEST
                )

            serializer.save()
            return Response(
                {"message": "Sent for Approval"},
                status=status.HTTP_201_CREATED
            )
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class CashInflowByProjectAPIView(APIView):
    """
    Retrieve all cash inflows associated with a specific project ID.
    """
    def get(self, request, pk):
        try:
            # Optimization: select_related fetches the project and core project in one query
            cash_inflows = CashInflow.objects.filter(
                project_id=pk
            ).select_related("project", "project__project")

            if not cash_inflows.exists():
                # Optional: Return empty list or 404 depending on your preference
                return Response([], status=status.HTTP_200_OK)

            serializer = CashInflowListSerializer(cash_inflows, many=True)
            return Response(serializer.data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": "Error Fetching Data", "detail": str(e)}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
 
def month_key(date):
    return date.strftime("%Y-%m")
 
class ProjectExcelFileListView(generics.ListAPIView):
    serializer_class = ExcelFileSerializer

    def get_queryset(self):
        project_id = self.kwargs.get('project_id')
        # Use select_related to join the Project table in a single query
        return ExcelFile.objects.filter(project_id=project_id).select_related('project')

class ProjectBoardListCreateAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]



    def get(self, request, project_id):
        purchase_project = get_object_or_404(PurchaseProject, id=project_id)
        
        capacity_name = purchase_project.project.project_name if purchase_project.project else ""
        code = purchase_project.project.co_no if purchase_project.project else ""
        pb_ct = ContentType.objects.get_for_model(ProjectBoard)

        all_pbs = ProjectBoard.objects.filter(project=purchase_project).prefetch_related('approvers', 'supplier_negotiations')
        project_budget = purchase_project.total_budget or 0
        
        total_spent = SupplierNegotiation.objects.filter(
            project_board__project=purchase_project,
            is_finalized=True
        ).aggregate(total=Sum('net_total'))['total'] or 0
        
        remaining_budget = float(project_budget) - float(total_spent)
        processed_pb_data = []

        # 1. Pre-fetch all relevant ContentTypes

        # 1. Pre-fetch ContentTypes for precision
        cts = {
            'pb': ContentType.objects.get_for_model(ProjectBoard),
            'comm': ContentType.objects.get_for_model(CommercialRequirement),
            'term': ContentType.objects.get_for_model(TermAssignment),
            'supp': ContentType.objects.get_for_model(Supplier),
            'supp_nego': ContentType.objects.get_for_model(SupplierNegotiation),
            'line_nego': ContentType.objects.get_for_model(LineItemNegotiation),
        }
        summary_counts = {
            "forms_under_review": 0,
            "drafts": 0,
            "rejected": 0,
            "approved": 0
        }
        for pb in all_pbs:
            # --- 1. Fetch Basic Approval/Nego State ---
            approval_records = ApprovalRejection.objects.filter(pb_id=pb)
            has_approval_started = approval_records.exists()
            
            total_required = approval_records.count()
            approved_count = approval_records.filter(status=ApprovalStatus.APPROVED).count()
            is_any_rejected = approval_records.filter(status=ApprovalStatus.REJECTED).exists()
            is_fully_approved = (total_required > 0 and approved_count == total_required)
            
            negotiations = pb.supplier_negotiations.all()
            is_any_nego_finalized = negotiations.filter(is_finalized=True).exists()

            # --- 2. STEP STATUS HELPER ---
            def get_step_info(step_label, related_objects_list):
                """
                related_objects_list: List of (ContentType, Object_ID) tuples.
                Example: [(cts['pb'], pb.id)]
                """
                # A. PRIORITY 1: CHECK FOR CHANGE REQUESTS
                for ct, obj_id in related_objects_list:
                    if obj_id:
                        cr = ChangeRequest.objects.filter(
                            content_type=ct, 
                            pending_data__project_board=pb.id, # Ensure CR is linked to the same PB
                            is_approved=False
                        ).select_related('requested_by').first()
                        
                        if cr:
                            status_text = f"Change Request Pending by {pb.project.project_manager.get_full_name() if cr.requested_by else 'Admin'}"
                            print(f"  [SUB-STEP] {step_label:25} -> {status_text}")
                            return {
                                "status": status_text,
                                "status_code": 2, # Warning code
                                "requested_by": cr.requested_by.get_full_name() if cr.requested_by else "system"
                            }

                # B. PRIORITY 2: CHECK FOR REJECTIONS
                model_key = step_label.lower().replace(" ", "")
                rej = approval_records.filter(content_type__model=model_key, status=ApprovalStatus.REJECTED).first()
                if rej:
                    return {"status": "Rejected","Action_by": rej.created_by.get_full_name() if rej.created_by else "N/A", "status_code": 2, "comment": rej.comments}

                # C. DATA PRESENCE
                primary_obj_exists = any(obj_id is not None for _, obj_id in related_objects_list)
                return {
                    "status": "Filled" if primary_obj_exists else "Not Filled", 
                    "status_code": 1 if primary_obj_exists else 0
                }

            # --- 3. GATHER IDs FOR CHILD TABLES ---
            comm_req = CommercialRequirement.objects.filter(project_board=pb).first()
            term_assign = TermAssignment.objects.filter(project_board=pb).first()
            # If term assignment exists, get linked supplier
            supp_id = term_assign.supplier_id if term_assign else None
            
            # Negotiation IDs (Supplier Nego + all Line Item Negos)
            supp_nego = negotiations.filter(is_finalized=True).first()
            line_nego_ids = list(LineItemNegotiation.objects.filter(project_board=pb).values_list('id', flat=True))
            if not has_approval_started:
                approval_status_label = "Not Filled"
                approval_status_code = 0
            elif is_any_rejected:
                approval_status_label = "Rejected"
                approval_status_code = 2
            elif is_fully_approved:
                approval_status_label = "Approved"
                approval_status_code = 1
            else:
                approval_status_label = "Pending"
                approval_status_code = 1
            # --- 4. ASSEMBLE ---
            step_status = {
                "basic_details": get_step_info("ProjectBoard", [(cts['pb'], pb.id)]),
                
                "commercial_requirements": get_step_info("CommercialRequirement", [(cts['comm'], comm_req.id if comm_req else None)]),
                
                "supplier": get_step_info("Supplier", [
                    (cts['term'], term_assign.id if term_assign else None),
                    (cts['supp'], supp_id)
                ]),
                
                "line_item": get_step_info("LineItemNegotiation", [
                    (cts['supp_nego'], supp_nego.id if supp_nego else None)
                ] + [(cts['line_nego'], lid) for lid in line_nego_ids]),
                
                "approval_request": {
                        "status": approval_status_label,
                        "status_code": approval_status_code,
                        "progress": f"{approved_count}/{total_required} total signs"
                    },            
            }




            # --- 5. MAIN VISUAL PB STATUS ---
            has_active_cr = any("Change Request" in str(v['status']) for v in step_status.values())

            if has_active_cr:
                pb_status = "Under Review: Change Request"
                summary_counts["forms_under_review"] += 1
                print("==================",summary_counts["forms_under_review"],summary_counts)

                
            elif not negotiations.exists() and not has_approval_started:
                pb_status = "Draft: Initial Data Entry"
                summary_counts["drafts"] += 1
                print("==================",summary_counts["drafts"],summary_counts)
                
            elif is_any_nego_finalized and not has_approval_started:
                pb_status = "Action Required: Ready for Submission"
                summary_counts["drafts"] += 1
                print("==================",summary_counts["drafts"],summary_counts)
                
            elif is_any_rejected:
                pb_status = "Rejected: Review Comments"
                summary_counts["rejected"] += 1
                print("==================",summary_counts["rejected"],summary_counts)
                
            elif is_fully_approved:
                pb_status = "Approved: Finalized"
                summary_counts["approved"] += 1
                print("==================",summary_counts["approved"],summary_counts)
                
            elif has_approval_started:
                pb_status = f"Pending: {approved_count}/{total_required} Approved"
                summary_counts["forms_under_review"] += 1
                print("==================",summary_counts["forms_under_review"],summary_counts)

            else:
                pb_status = "Unknown Status"
                print("==================",summary_counts["forms_under_review"])


            print(f"  >>> FINAL MAIN STATUS: {pb_status}\n")

            # --- 6. SERIALIZE ---
            item = ProjectBoardSummarySerializer(pb).data
            item.update({
                'status': pb_status,
                'step_status': step_status,
                'has_active_cr': has_active_cr
            })
            processed_pb_data.append(item)
        # 7. Standalone Drafts
        standalone_drafts = Draft.objects.filter(
            content_type=pb_ct, is_approved=False,
            pending_data__project=capacity_name, pending_data__cost_center=code,
        )
        
        draft_data = []
        for draft in standalone_drafts:
            summary_counts["drafts"] += 1     
            print("==================",summary_counts["drafts"],summary_counts)

            content = draft.pending_data
            if isinstance(content, dict):
                item = content.copy()
                item.update({
                    'is_draft': True,
                    'status': '🟠 Draft',
                    'draft_id': draft.id,
                    'step_status': None
                })
                draft_data.append(item)
        print("==================",summary_counts)

        return Response({
            "subject": capacity_name, "code": code,
            "project_budget": project_budget, "total_spent": total_spent,"summary_counts": summary_counts,
            "remaining_budget": remaining_budget, "results": processed_pb_data + draft_data
        }, status=status.HTTP_200_OK)


    def post(self, request):
        serializer = ProjectBoardCreateUpdateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # 1. Create the ProjectBoard
                project_board = serializer.save(created_by=request.user)

                # 2. Automatically create an approved ChangeRequest for creation audit
                content_type = ContentType.objects.get_for_model(ProjectBoard)
  
                
                ChangeRequest.objects.create(
                    content_type=content_type,
                    object_id=project_board.id,
                    # Store the full raw incoming data (recommended for traceability)
                    pending_data=request.data.copy(),  # .copy() to avoid reference issues
                    requested_by=request.user,
                    is_approved=True,                    # creation is considered approved
                )

                # 3. Return the newly created object
                return Response(
                    ProjectBoardReadSerializer(project_board).data,
                    status=status.HTTP_201_CREATED
                )

        except Exception as e:
            return Response(
                {"error": "Failed to create ProjectBoard", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


from django.db.models import Sum, Count, Q
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.contenttypes.models import ContentType

from django.db.models import Sum, Count, F
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.contrib.contenttypes.models import ContentType



class BuyerProcurementDashboard(APIView):
    def get(self, request):
        buyer = request.user
        pb_ct = ContentType.objects.get_for_model(ProjectBoard)
        
        all_buyer_pbs = ProjectBoard.objects.filter(created_by=buyer).prefetch_related(
            'approvers', 'supplier_negotiations', 'project__project_manager'
        )

        # 1. Initialize lists and counters
        total_pb_count = all_buyer_pbs.count()
        approved_pbs_count = 0
        total_commodity_budget = 0
        total_actual_expense = 0
        
        # Consistent naming: 'rejected_list' and 'under_review_list'
        rejected_list = []
        under_review_list = []

        for pb in all_buyer_pbs:
            # --- Financials ---
            comm_budget = float(pb.commodity_budget or 0)
            total_commodity_budget += comm_budget
            
            actual_spent = pb.supplier_negotiations.filter(is_finalized=True).aggregate(
                total=Sum('net_total'))['total'] or 0
            total_actual_expense += float(actual_spent)

            # --- Approvals/Rejections ---
            approval_records = ApprovalRejection.objects.filter(pb_id=pb)
            total_req = approval_records.count()
            approved_count = approval_records.filter(status="APPROVED").count()
            rejections = approval_records.filter(status="REJECTED").select_related('created_by', 'content_type')

            if total_req > 0 and approved_count == total_req:
                approved_pbs_count += 1
            
            if rejections.exists():
                reasons = [{
                    "table": r.content_type.model,
                    "rejected_by": r.created_by.get_full_name() if r.created_by else "Approver",
                    "comment": r.comments
                } for r in rejections]
                
                rejected_list.append({
                    "document_no": pb.document_no,
                    "reasons": reasons
                })

            # --- Change Requests (Under Review by PM) ---
            crs = ChangeRequest.objects.filter(content_type=pb_ct, object_id=pb.id, is_approved=False)
            if crs.exists():
                under_review_list.append({
                    "document_no": pb.document_no,
                    "pending_at": pb.project.project_manager.get_full_name() if pb.project.project_manager else "PM",
                    "changes": [{"reason": getattr(cr,'reason','No reason provided'), "date": cr.created_at} for cr in crs]
                })

        # --- Variance Calculation ---
        budget_variance = total_commodity_budget - total_actual_expense

        # 2. Return Response using the corrected names
        return Response({
            "buyer_summary": {
                "total_pbs": total_pb_count,
                "approved_pbs": approved_pbs_count,
                "rejected_pbs_count": len(rejected_list),
                "under_review_count": len(under_review_list),
            },
            "financial_summary": {
                "total_budget_allocated": total_commodity_budget,
                "total_actual_expense": total_actual_expense,
                "variance": budget_variance,
                "budget_status": "UNDER_BUDGET" if budget_variance >= 0 else "OVER_BUDGET"
            },
            "action_items": {
                "rejected_details": rejected_list,
                "under_review_details": under_review_list
            }
        }, status=status.HTTP_200_OK)

class ProjectBoardDetailAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request, pb_id):
        # 1. Fetch the ProjectBoard
        pb = get_object_or_404(ProjectBoard, id=pb_id)
        
        # 2. Get ContentType
        content_type = ContentType.objects.get_for_model(pb)
        
        # 3. Fetch ALL approval/rejection records for this board
        # Ordered by newest first so the latest comment is at the top
        approval_history = ApprovalRejection.objects.filter(
            pb_id=pb, 
            content_type=content_type,
            status__in=[ApprovalStatus.REJECTED]
        ).order_by('-created_at')

        # 4. Serialize the ProjectBoard data
        data = ProjectBoardReadSerializer(pb).data
        
        # 5. Serialize and attach the history list
        data['approval_history'] = ApprovalRejectionSerializer(approval_history, many=True).data

        return Response(data)

    # def patch(self, request, pb_id):
    #     pb = get_object_or_404(ProjectBoard, id=pb_id)
        
    #     # 1. Validate data against the model's rules
    #     serializer = ProjectBoardCreateUpdateSerializer(pb, data=request.data, partial=True)
        
    #     if serializer.is_valid():
    #         # 2. Get ContentType for the generic foreign key
    #         content_type = ContentType.objects.get_for_model(pb)
            
    #         existing_request = ChangeRequest.objects.filter(
    #             content_type=content_type,
    #             object_id=pb.id,
    #             is_approved=False
    #         ).exists()

    #         if existing_request:
    #             return Response(
    #                 {"error": "A pending change request already exists for this record. Please wait for approval or rejection."},
    #                 status=status.HTTP_400_BAD_REQUEST
    #             )
            
    #         change_request = ChangeRequest.objects.create(
    #             content_type=content_type,
    #             object_id=pb.id,
    #             pending_data=request.data, # Store the raw updated fields
    #             requested_by=request.user,
    #             is_approved=False
    #         )
            
    #         return Response({
    #             "message": "Update intercepted. Change request created.",
    #             "change_request_id": change_request.id,
    #             "status": "Pending Approval"
    #         }, status=status.HTTP_202_ACCEPTED)
            
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
    def patch(self, request, pb_id):
        print(f"\n--- PATCH START: PB ID {pb_id} ---")
        pb = get_object_or_404(ProjectBoard, id=pb_id)
        serializer = ProjectBoardCreateUpdateSerializer(pb, data=request.data, partial=True)
        
        if serializer.is_valid():
            content_type = ContentType.objects.get_for_model(pb)
            
            # --- GLOBAL LOCK CHECK ---
            # If ANY change request is pending for this record, block ALL updates.
            existing_request = ChangeRequest.objects.filter(
                content_type=content_type,
                object_id=pb.id,
                is_approved=False
            ).exists()

            if existing_request:
                print(f"DEBUG: Blocked - A pending change request exists for {content_type.model}")
                return Response(
                    {"error": "This record is currently locked. A pending change request must be approved or rejected first."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )

            # 1. Check the status of THIS specific model
            current_approval = ApprovalRejection.objects.filter(
                pb_id=pb_id, 
                content_type=content_type,
                status__in=[ApprovalStatus.REJECTED]
            ).first()

            # CASE 1: This specific record is REJECTED (and no ChangeRequest exists)
            if current_approval and current_approval.status == ApprovalStatus.REJECTED:
                print("DEBUG: Entered REJECTED Path (Direct Update)")
                with transaction.atomic():
                    serializer.save()
                    
                    # Check for other rejections
                    other_rejected_exists = ApprovalRejection.objects.filter(
                        pb_id=pb_id, 
                        status=ApprovalStatus.REJECTED
                    ).exclude(id=current_approval.id).exists()

                    if other_rejected_exists:
                        current_approval.status = ApprovalStatus.PENDING
                        current_approval.save()
                        message = "Section updated. Other rejections still exist."
                    else:
                        # Final Fix: Full Board Reset
                        ApprovalRejection.objects.filter(pb_id=pb_id).update(
                            status=ApprovalStatus.PENDING,
                            comments=None,
                            created_at=timezone.now()
                        )
                        message = "Final rejection fixed. All board records reset to Pending."

                return Response({"message": message, "data": serializer.data}, status=status.HTTP_200_OK)

            # CASE 2: Approved/Pending status -> Create new ChangeRequest
            else:
                print("DEBUG: Creating new ChangeRequest")
                change_request = ChangeRequest.objects.create(
                    content_type=content_type,
                    object_id=pb.id,
                    pending_data=request.data,
                    requested_by=request.user,
                    is_approved=False
                )
                
                return Response({
                    "message": "Change request created.",
                    "change_request_id": change_request.id
                }, status=status.HTTP_202_ACCEPTED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class CommercialRequirementListCreateAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request, pb_id):
        pbs = CommercialRequirement.objects.filter(project_board=pb_id)
        return Response(CommercialRequirementSummarySerializer(pbs, many=True).data)

    # def post(self, request):
    #     serializer = CommercialRequirementCreateUpdateSerializer(data=request.data)
    #     if serializer.is_valid():
    #         instance = serializer.save(created_by=request.user)
    #         return Response(CommercialRequirementCreateUpdateSerializer(instance).data, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        serializer = CommercialRequirementCreateUpdateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # 1. Create the CommercialRequirement
                requirement = serializer.save(created_by=request.user)

                # 2. Automatically create an approved ChangeRequest (audit/history)
                content_type = ContentType.objects.get_for_model(CommercialRequirement)
                
                ChangeRequest.objects.create(
                    content_type=content_type,
                    object_id=requirement.id,
                    # Recommended: store raw incoming data for full traceability
                    pending_data=request.data.copy(),  # .copy() prevents reference issues
                    # Alternative: use validated data (cleaner but less raw)
                    # pending_data=serializer.validated_data.copy(),
                    requested_by=request.user,
                    is_approved=True,                    # creation is auto-approved
                )

                # 3. Return the created object
                return Response(
                    CommercialRequirementCreateUpdateSerializer(requirement).data,
                    status=status.HTTP_201_CREATED
                )

        except Exception as e:
            return Response(
                {"error": "Failed to create CommercialRequirement", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class CommercialRequirementDetailAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    # def get(self, request, pb_id):
    #     pb = get_object_or_404(CommercialRequirement, project_board=pb_id)
    #     return Response(CommercialRequirementSummarySerializer(pb).data)

    def get(self, request, pb_id):
        # 1. Prepare ContentTypes (we'll need this for the history search)
        ct = ContentType.objects.get_for_model(CommercialRequirement)
        
        # 2. Fetch Rejection History (Multiple records possible)
        # We filter by pb_id (ProjectBoard) and the ContentType of CommercialRequirement
        rejections = ApprovalRejection.objects.filter(
            pb_id=pb_id,
            content_type=ct,
            status__in=[ApprovalStatus.REJECTED]
        ).order_by('-created_at')
        history_data = ApprovalRejectionSerializer(rejections, many=True).data
        
        # Map the history to a list of dictionaries
            # history_data = [
            #     {
            #         "status": r.status,
            #         "comments": r.comments,
            #         "date": r.created_at
            #     } for r in rejections
            # ]

        # 3. Try to find the actual record first
        pb_record = CommercialRequirement.objects.filter(project_board=pb_id).first()
        
        if pb_record:
            data = CommercialRequirementSummarySerializer(pb_record).data
            data['approval_history'] = history_data
            return Response(data)

        # 4. If not found, look in the Draft/ChangeRequest table
        draft = Draft.objects.filter(
            content_type=ct,
            requested_by=request.user,
            is_approved=False
        ).filter(pending_data__project_board=pb_id).first()

        if draft:
            return Response({
                "draft_id": draft.id,
                "is_draft": True,
                **draft.pending_data,
                "approval_history": history_data # Attach history even to drafts
            })

        # 5. If history exists but no record/draft, we might still want to show the history
        if history_data:
            return Response({
                "message": "No active record or draft, but history exists.",
                "approval_history": history_data
            })

        # 6. If absolutely nothing exists
        return Response(
            {"error": "No record, draft, or history found for this Project Board."}, 
            status=status.HTTP_404_NOT_FOUND
        )

    def patch(self, request, pb_id):
        pb = get_object_or_404(CommercialRequirement, project_board=pb_id)
        serializer = CommercialRequirementCreateUpdateSerializer(pb, data=request.data, partial=True)
        
        if serializer.is_valid():
            content_type = ContentType.objects.get_for_model(CommercialRequirement)
            existing_request = ChangeRequest.objects.filter(
                content_type=content_type,
                object_id=pb.id,
                is_approved=False
            ).exists()

            if existing_request:
                print(f"DEBUG: Blocked - A pending change request exists for {content_type.model}")
                return Response(
                    {"error": "This record is currently locked. A pending change request must be approved or rejected first."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )            
            # 1. Check if THIS specific CommercialRequirement record is REJECTED
            current_approval = ApprovalRejection.objects.filter(
                pb_id=pb_id, # Or pb.project_board_id depending on your field name
                content_type=content_type,
                status__in=[ApprovalStatus.REJECTED]
            ).first()

            # CASE 1: Fixing a Rejected record
            if current_approval and current_approval.status == ApprovalStatus.REJECTED:
                with transaction.atomic():
                    # Direct Save
                    serializer.save()
                    
                    # Check for other rejections for the same Project Board
                    other_rejected_exists = ApprovalRejection.objects.filter(
                        pb_id=pb_id, 
                        status=ApprovalStatus.REJECTED
                    ).exclude(id=current_approval.id).exists()

                    if other_rejected_exists:
                        # Just move this one to Pending
                        current_approval.status = ApprovalStatus.PENDING
                        current_approval.save()
                        message = "Commercial Requirement updated. Other rejections still pending fix."
                    else:
                        # LAST REJECTION FIXED: Full Reset
                        ApprovalRejection.objects.filter(pb_id=pb_id).update(
                            status=ApprovalStatus.PENDING,
                            comments=None,
                            created_at=timezone.now()
                        )
                        message = "Final rejection fixed. All board records reset to Pending with comments cleared."

                return Response({
                    "message": message,
                    "data": serializer.data
                }, status=status.HTTP_200_OK)

            # CASE 2: Normal Update (ChangeRequest Workflow)
            else:
                existing_request = ChangeRequest.objects.filter(
                    content_type=content_type,
                    object_id=pb.id,
                    is_approved=False
                ).exists()

                if existing_request:
                    return Response(
                        {"error": "A pending change request already exists."},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                change_request = ChangeRequest.objects.create(
                    content_type=content_type,
                    object_id=pb.id,
                    pending_data=request.data,
                    requested_by=request.user,
                    is_approved=False
                )
                
                return Response({
                    "message": "Change request created for approval.",
                    "change_request_id": change_request.id
                }, status=status.HTTP_202_ACCEPTED)
                
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class SupplierListCreateAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request):
        pbs = Supplier.objects.all()
        return Response(SupplierSummarySerializer(pbs, many=True).data)



    # def post(self, request):
    #     serializer = SupplierCreateUpdateSerializer(data=request.data)
    #     if serializer.is_valid():
    #         instance = serializer.save(created_by=request.user)
    #         return Response(SupplierCreateUpdateSerializer(instance).data, status=status.HTTP_201_CREATED)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def post(self, request):
        serializer = SupplierCreateUpdateSerializer(data=request.data)
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        try:
            with transaction.atomic():
                # 1. Create the Supplier record
                supplier = serializer.save(created_by=request.user)

                # 2. Automatically create an approved ChangeRequest (audit/history)
                content_type = ContentType.objects.get_for_model(Supplier)
                
                ChangeRequest.objects.create(
                    content_type=content_type,
                    object_id=supplier.id,
                    pending_data=request.data.copy(),  
                    requested_by=request.user,
                    is_approved=True,
                )

                # 3. Return the created supplier
                return Response(
                    SupplierCreateUpdateSerializer(supplier).data,
                    status=status.HTTP_201_CREATED
                )

        except Exception as e:
            return Response(
                {"error": "Failed to create Supplier", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
import pprint  # Import at the top of your file
class SupplierDetailAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]


    def get(self, request, pb_id):
        # 1. Fetch Real Suppliers
        pbs = Supplier.objects.filter(project_board_id=pb_id)
        real_data = SupplierSummarySerializer(pbs, many=True).data

        # 2. Total Terms
        total_system_terms = Term.objects.count()
        
        # 3. Aggregate Stats
        assignment_stats = TermAssignment.objects.filter(
            project_board_id=pb_id  # This must be the ID for board 33
        ).values('supplier_id').annotate(
            count_expected=Count('term_id', filter=Q(status_type='Expected'), distinct=True),
            count_agreed=Count('term_id', filter=Q(status_type='Agreed'), distinct=True)
        )

        # Build the lookup
        stats_lookup = {
            entry['supplier_id']: {
                'expected': entry['count_expected'],
                'agreed': entry['count_agreed']
            } for entry in assignment_stats
        }

        # --- ADD THIS DEBUG LINE TO SEE RAW QUERY ---
        print(f"DEBUG: SQL Query being run: {assignment_stats.query}")

        # --- PPRINT THE LOOKUP TABLE ---
        print("\n" + "="*60)
        print(f"DEBUG START: PROJECT BOARD {pb_id}")
        print(f"Total Terms defined in DB: {total_system_terms}")
        print("Stats Lookup (Term Assignments found in DB):")
        pprint.pprint(stats_lookup)
        print("="*60 + "\n")

        # 4. Enrich Real Records
        for item in real_data:
            supplier_id = item['id']
            counts = stats_lookup.get(supplier_id, {'expected': 0, 'agreed': 0})
            
            is_expected_complete = (total_system_terms > 0 and counts['expected'] >= total_system_terms)
            is_agreed_complete = (total_system_terms > 0 and counts['agreed'] >= total_system_terms)

            # --- PPRINT INDIVIDUAL SUPPLIER CALCULATION ---
            print(f"Processing Supplier: {item.get('supplier_name')} (ID: {supplier_id})")
            print(f"  > Expected: {counts['expected']}/{total_system_terms} -> Result: {is_expected_complete}")
            print(f"  > Agreed:   {counts['agreed']}/{total_system_terms} -> Result: {is_agreed_complete}")
            print("-" * 30)

            item.update({
                'is_draft': False,
                'expected_filled': is_expected_complete,
                'agreed_filled': is_agreed_complete,
                'has_terms': is_expected_complete and is_agreed_complete
            })
        # --- MOVED OUTSIDE THE LOOP ---
        # 5. Setup Draft search (Once per request)
        pb_ct = ContentType.objects.get_for_model(Supplier)
        draft_queryset = Draft.objects.filter(
            content_type=pb_ct, 
            is_approved=False
        ).filter(
            Q(pending_data__form__project_board=str(pb_id)) | 
            Q(pending_data__form__project_board=int(pb_id)) |
            Q(pending_data__project_board=str(pb_id)) |
            Q(pending_data__project_board=int(pb_id))
        )

        draft_data = []
        for draft in draft_queryset:
            content = draft.pending_data.get('form', draft.pending_data)
            if isinstance(content, dict):
                item = content.copy()
                item['is_draft'] = True
                item['draft_id'] = draft.id
                item['expected_filled'] = False
                item['agreed_filled'] = False
                item['has_terms'] = False 
                draft_data.append(item)

        return Response(real_data + draft_data, status=status.HTTP_200_OK)

    def delete(self, request, pb_id):
        # We use pb_id here because your URL route likely maps it to the Supplier's ID
        supplier = get_object_or_404(Supplier, id=pb_id)
        
        # Optional: Check if there are pending change requests and clean them up
        # content_type = ContentType.objects.get_for_model(Supplier)
        # ChangeRequest.objects.filter(content_type=content_type, object_id=supplier.id).delete()

        supplier.delete()
        
        return Response(
            {"message": f"Supplier {pb_id} deleted successfully."}, 
            status=status.HTTP_204_NO_CONTENT
        )

    def patch(self, request, pb_id):
        pb = get_object_or_404(Supplier, id=pb_id)
        
        serializer = SupplierCreateUpdateSerializer(pb, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        EMPTY_VALUES = [None, "", "NA", "none", "None"]
        required_fields = [
            'supplier_name', 'quotation_validity_before', 'quotation_validity_after',
            'energy_efficiency_before', 'energy_efficiency_after', 
            'risk_category_before', 'risk_category_after'
        ]

        # 1. Check if the record is currently complete (before this update)
        is_already_complete = all([getattr(pb, f) not in EMPTY_VALUES for f in required_fields])

        with transaction.atomic():
            if not is_already_complete:
                # SCENARIO A: The record is still in the "Initial Fill" stage.
                # We save all incoming valid data directly to the table.
                for field, value in serializer.validated_data.items():
                    setattr(pb, field, value)
                pb.save()
                
                return Response({
                    "message": "Record updated directly (Initial fill stage).",
                    "data": SupplierCreateUpdateSerializer(pb).data
                }, status=status.HTTP_200_OK)

            else:
                # SCENARIO B: Record is already full. Check for actual changes.
                # Compare current DB values with incoming validated data.
                changes = {
                    field: value for field, value in serializer.validated_data.items() 
                    if getattr(pb, field) != value
                }

                if changes:
                    content_type = ContentType.objects.get_for_model(pb)
                    
                    # Prevent multiple pending requests for the same object
                    if ChangeRequest.objects.filter(content_type=content_type, object_id=pb.id, is_approved=False).exists():
                        return Response(
                            {"error": "A pending change request already exists for this supplier."}, 
                            status=status.HTTP_400_BAD_REQUEST
                        )

                    # Create ChangeRequest with the FULL data (request.data) as requested
                    cr = ChangeRequest.objects.create(
                        content_type=content_type,
                        object_id=pb.id,
                        pending_data=request.data,  # Captures the full state of the form
                        requested_by=request.user
                    )
                    
                    return Response({
                        "message": "Change request created (Record was already complete).",
                        "changed_fields": list(changes.keys()),
                        "cr_id": cr.id
                    }, status=status.HTTP_202_ACCEPTED)

        # If no changes were detected
        return Response(SupplierCreateUpdateSerializer(pb).data, status=status.HTTP_200_OK)




class LineItemListCreateAPIView(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    def get(self, request):
        pbs = Supplier.objects.all()
        return Response(SupplierSummarySerializer(pbs, many=True).data)


    def post(self, request):
        serializer = LineItemSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        validated_data = serializer.validated_data
        # Extract negotiations for creation logic
        negotiations_data = validated_data.pop('negotiations', [])
        user = request.user if request.user.is_authenticated else None

        # Helper to handle Decimal/Date serialization in JSON
        def serialize_data(data):
            return json.loads(json.dumps(data, cls=DjangoJSONEncoder))

        try:
            with transaction.atomic():
                # 1. Create the LineItem record
                line_item = LineItem.objects.create(
                    created_by=user,
                    **validated_data
                )

                # 2. Create the Negotiation records
                negotiation_ids = []
                for neg_item in negotiations_data:
                    negotiation = LineItemNegotiation.objects.create(
                        project_board=line_item.project_board,
                        line_item=line_item,
                        created_by=user,
                        **neg_item
                    )
                    negotiation_ids.append(negotiation.id)

                # 3. Create ONE SINGLE ChangeRequest for everything
                # This stores the FULL 'request.data' (original payload) in one place
                li_ct = ContentType.objects.get_for_model(LineItem)
                ChangeRequest.objects.create(
                    content_type=li_ct,
                    object_id=line_item.id,
                    pending_data=serialize_data(request.data), # Store the whole payload
                    requested_by=request.user,
                    is_approved=True  # Marking as approved since it's the initial creation
                )

                return Response(
                    {
                        "message": "Line item and negotiations created with a single audit log.",
                        "line_item_id": line_item.id,
                        "negotiation_ids": negotiation_ids
                    },
                    status=status.HTTP_201_CREATED
                )

        except Exception as e:
            return Response(
                {"error": "Failed to create LineItem", "detail": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


class ExcelUploadView(APIView):
    parser_classes = (MultiPartParser, FormParser)

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')
        
        if not file_obj:
            return Response({"error": "No file provided"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Call your processing logic
            df = self.handle_excel_data(file_obj)
            
            # Convert DataFrame to JSON-friendly dictionary
            data = df.to_dict(orient='records')
            
            return Response({
                "message": "File processed successfully",
                "total_rows": len(data),
                "columns": list(df.columns),
                "data": data
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)

    def handle_excel_data(self, file):
        # --- Your existing logic logic ---
        xl = pd.ExcelFile(file)
        target_sheet = "Procurement Board"
        
        sheets_lower = {s.lower().strip(): s for s in xl.sheet_names}
        if target_sheet.lower() in sheets_lower:
            target_sheet = sheets_lower[target_sheet.lower()]
        else:
            raise Exception(f"Sheet '{target_sheet}' not found.")

        header_data = xl.parse(target_sheet, nrows=3, header=None)
        
        # Row 0 ffill: ['Budgets', NaN, 'ABC', NaN...] -> ['Budgets', 'Budgets', 'ABC', 'ABC'...]
        parent_headers = header_data.iloc[0].ffill().fillna("")
        child_headers = header_data.iloc[2].fillna("")

        final_columns = []
        for p, c in zip(parent_headers, child_headers):
            p_str = str(p).strip().replace(" ", "_").lower()
            c_str = str(c).strip().replace(" ", "_").lower()
            
            if p_str and p_str != c_str:
                combined_name = f"{p_str}_{c_str}"
            else:
                combined_name = c_str
            
            combined_name = combined_name.replace("\n", "_").replace("\"", "")
            final_columns.append(combined_name)

        df = xl.parse(target_sheet, skiprows=3, header=None)
        df.columns = final_columns
        df = df.dropna(how="all")
        df = df.loc[:, ~df.columns.str.contains("^unnamed|^$", na=False)]
        
        # Clean up column names for JSON (remove dots/spaces)
        df.columns = [c.strip().replace(" ", "_").replace(".", "") for c in df.columns]

        return df


class LineItemDetailAPIViewAll(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    def get(self, request, pb_id):
        # 1. Fetch the Project Board and Project
        try:
            project_board = ProjectBoard.objects.select_related('project').get(id=pb_id)
            project = project_board.project
        except ProjectBoard.DoesNotExist:
            return Response({"error": "Project Board not found"}, status=status.HTTP_404_NOT_FOUND)

        # 2. NEW: Fetch Rejection History for Line Items
        # We use the ContentType of LineItem to get history specific to this section
        line_item_ct = ContentType.objects.get_for_model(LineItemNegotiation)  # Assuming negotiations are the main source of rejections for line items
        rejections = ApprovalRejection.objects.filter(
            pb_id=pb_id, 
            content_type=line_item_ct,
            status__in=[ApprovalStatus.REJECTED]            
        ).order_by('-created_at')

        approval_history = ApprovalRejectionSerializer(rejections, many=True).data

        # 3. Calculate Budget
        total_budget = project_board.commodity_budget or 0
        total_used = LineItem.objects.filter(project_board=project_board).aggregate(
            total=Sum('budget')
        )['total'] or 0
        
        remaining_budget = total_budget - total_used

        # 4. Get the Line Items
        line_items = LineItem.objects.filter(project_board=project_board)
        serializer = LineItemSerializer(line_items, many=True)

        # 5. Final Response Structure
        return Response({
            "budget_remaining_project_wide": remaining_budget,
            "budget_comodity": total_budget,
            "line_items": serializer.data,
            "approval_history": approval_history  # Added history here
        }, status=status.HTTP_200_OK)



class LineItemDetailAPIView(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    # def get(self, request, li_id):
    #     # Logic for get_line_item_by_id
    #     line_item = LineItem.objects.filter(id=li_id)
    #     serializer = LineItemSerializer(line_item, many=True)
    #     return Response(serializer.data, status=status.HTTP_200_OK)

    def get(self, request, li_id):
        # 1. Prepare ContentType for LineItem
        ct = ContentType.objects.get_for_model(LineItemNegotiation)
        
        # 2. Fetch Rejection History
        # Note: Ensure your ApprovalRejection model has a way to link to the specific LineItem (li_id)
        # usually via object_id or a specific line_item field.
        rejections = ApprovalRejection.objects.filter(
            content_type=ct,
            object_id=li_id, # Link directly to this specific line item
            status=ApprovalStatus.REJECTED
        ).order_by('-created_at')
        
        history_data = ApprovalRejectionSerializer(rejections, many=True).data

        # 3. Try to find the Actual Record (Live Data)
        line_item = LineItem.objects.filter(id=li_id).first()
        
        if line_item:
            serializer = LineItemSerializer(line_item)
            data = serializer.data
            # Attach history directly to the object data
            data['rejection_history'] = history_data
            return Response(data, status=status.HTTP_200_OK)

        # 4. If not found, look in the ChangeRequest/Draft table
 
        if draft:
            return Response({
                "draft_id": draft.id,
                "is_draft": True,
                "approval_history": history_data
            }, status=status.HTTP_200_OK)

        # 5. If only History exists
        if history_data:
            return Response({
                "message": "Line item record not found, but rejection history exists.",
                "approval_history": history_data
            }, status=status.HTTP_200_OK)

        # 6. Absolute 404
        return Response(
            {"error": "No record, draft, or history found for this Line Item."}, 
            status=status.HTTP_404_NOT_FOUND
        )        
    def patch(self, request, li_id):
        line_item = get_object_or_404(LineItem, id=li_id)
        incoming_data = request.data
        incoming_negotiations = incoming_data.get('negotiations', [])

        # Fields that should be calculated by the model save() method, not input by user
        calculated_fields = [
            'total_price_before', 
            'total_price_after', 
            'discount_overall_net', 
            'discount_overall_percantage'
        ]

        try:
            with transaction.atomic():
                # --- 1. Update Line Item Main Fields ---
                for field, value in incoming_data.items():
                    if field == 'negotiations':
                        continue
                    
                    # Map foreign key fields to their _id attribute
                    target_attr = f"{field}_id" if field in ['project_board', 'created_by', 'budget_enterd_by'] else field
                    
                    if hasattr(line_item, target_attr):
                        setattr(line_item, target_attr, value)
                
                line_item.save()

                # --- 2. Update Line Item Negotiations ---
                for neg_data in incoming_negotiations:
                    neg_id = neg_data.get('id')
                    supplier_id = neg_data.get('supplier')

                    neg_obj = None

                    # Strategy: Find existing record by ID, or fallback to Supplier + LineItem
                    if neg_id:
                        neg_obj = LineItemNegotiation.objects.filter(id=neg_id).first()
                    elif supplier_id:
                        neg_obj = LineItemNegotiation.objects.filter(
                            line_item=line_item, 
                            supplier_id=supplier_id
                        ).first()

                    # Skip if we can't identify which negotiation to update
                    if not neg_obj:
                        continue

                    for nf, nv in neg_data.items():
                        # Skip ID and calculated fields to prevent overwriting logic
                        if nf == 'id' or nf in calculated_fields:
                            continue

                        # Handle Foreign Key field names
                        if nf in ['project_board', 'supplier', 'line_item', 'finalized_by']:
                            target_nf = f"{nf}_id"
                        else:
                            target_nf = nf

                        if hasattr(neg_obj, target_nf):
                            # Ensure empty strings for numeric fields are treated as None or 0
                            if nv == '' and any(x in nf for x in ['price', 'total', 'discount']):
                                nv = Decimal('0.00')
                            
                            setattr(neg_obj, target_nf, nv)

                    # Saving the negotiation triggers model-level calculations
                    neg_obj.save()

            return Response({"message": "Line item and negotiations updated successfully."}, status=status.HTTP_200_OK)

        except Exception as e:
            # Captures specific database or logic errors for debugging
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)


class LineItemDetailAPIView1(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    def get(self, request, li_id):
        # Logic for get_line_item_by_id
        line_item = LineItem.objects.filter(id=li_id)
        serializer = LineItemSerializer(line_item, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, li_id):
        line_item = get_object_or_404(LineItem, id=li_id)
        incoming_data = request.data
        incoming_negotiations = incoming_data.get('negotiations', [])

        def is_empty(val):
            return val in [None, 0, '', Decimal('0.00'), '0.00', '0']

        def is_already_full(li_obj):
            critical_li = [li_obj.budget, li_obj.target_price, li_obj.quantity]
            if any(is_empty(v) for v in critical_li):
                return False
            negs = li_obj.negotiations.all()
            if not negs.exists():
                return False
            for n in negs:
                if is_empty(n.unit_price_after) or is_empty(n.total_price_after):
                    return False
            return True

        record_was_full = is_already_full(line_item)

        try:
            with transaction.atomic():
                if record_was_full:
                    # --- SCENARIO A: CHANGE REQUEST ---
                    li_ct = ContentType.objects.get_for_model(LineItem)
                    if ChangeRequest.objects.filter(content_type=li_ct, object_id=line_item.id, is_approved=False).exists():
                        return Response({"error": "A pending change request already exists."}, status=status.HTTP_400_BAD_REQUEST)

                    ChangeRequest.objects.create(
                        content_type=li_ct,
                        object_id=line_item.id,
                        pending_data=incoming_data,
                        requested_by=request.user
                    )
                    return Response({"message": "Record full. Change request created."}, status=status.HTTP_202_ACCEPTED)

                else:
                    # --- SCENARIO B: DIRECT PATCH ---
                    for field, value in incoming_data.items():
                        if field == 'negotiations': continue
                        target_attr = f"{field}_id" if field in ['project_board', 'created_by', 'budget_enterd_by'] else field
                        if hasattr(line_item, target_attr):
                            setattr(line_item, target_attr, value)
                    line_item.save()

                    # Update Negotiations
                    for neg_data in incoming_negotiations:
                        neg_id = neg_data.get('id')
                        if neg_id:
                            # CRITICAL: Added select_related to ensure quantity is available for math
                            neg_obj = LineItemNegotiation.objects.select_related('line_item').get(id=neg_id)
                            
                            # CRITICAL: List of fields to NOT overwrite with frontend data
                            calculated_fields = ['total_price_before', 'total_price_after', 'discount_overall_net', 'discount_overall_percantage']

                            for nf, nv in neg_data.items():
                                # SKIP fields that the model should calculate
                                if nf in calculated_fields:
                                    continue

                                target_nf = f"{nf}_id" if nf in ['project_board', 'supplier', 'line_item', 'finalized_by'] else nf
                                if hasattr(neg_obj, target_nf):
                                    setattr(neg_obj, target_nf, nv)
                            
                            # Now this call actually triggers the calculations properly
                            neg_obj.save()

                    return Response({"message": "Initial data filled directly."}, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"message": "Supplier finalized with aggregated totals."}, status=status.HTTP_200_OK)

# class SupplierNegotiationSummaryAPIView(APIView):

#     def get_calculations(self, pb_id, supplier_id):
#         aggregates = LineItemNegotiation.objects.filter(
#             project_board_id=pb_id,
#             supplier_id=supplier_id
#         ).aggregate(
#             sum_gross=Sum('total_price_before'),
#             sum_net=Sum('total_price_after'),
#             sum_discount_net=Sum('discount_overall_net')
#         )

#         gross = aggregates['sum_gross'] or Decimal('0.00')
#         net = aggregates['sum_net'] or Decimal('0.00')
#         discount_net = aggregates['sum_discount_net'] or Decimal('0.00')
#         percentage = (discount_net / gross * 100) if gross > 0 else Decimal('0.00')

#         return {
#             "gross_total": gross,
#             "net_total": net,
#             "discount_overall_net": discount_net,
#             "discount_overall_percentage": round(percentage, 2)
#         }

#     def get(self, request, pb_id, supplier_id):
#         # 1. Live calculations
#         stats = self.get_calculations(pb_id, supplier_id)
        
#         # 2. Database record
#         header = SupplierNegotiation.objects.filter(
#             project_board_id=pb_id, 
#             supplier_id=supplier_id
#         ).first()
        
#         # 3. Supplier context
#         supplier = get_object_or_404(Supplier, id=supplier_id)
#         terms_exist = TermAssignment.objects.filter(
#             project_board_id=pb_id,
#             supplier_id=supplier_id
#         ).exists()
#         # Unified Response Structure
#         return Response({
#             "id": header.id if header else None,
#             "project_board_id": pb_id,
#             "supplier_id": supplier_id,
#             "supplier_name": supplier.supplier_name,
#             "is_finalized": header.is_finalized if header else False,
#             "has_terms_assigned": terms_exist,  # <--- New field for UI logic
#             "finalized_details": {
#                 "finalized_at": header.finalized_at if header else None,
#                 "finalized_by": header.finalized_by.username if header and header.finalized_by else None,
#                 "comment": header.finalized_comment if header else ""
#             },
#             "overall_negotiation": stats 
#         }, status=status.HTTP_200_OK)

#     @transaction.atomic
#     def post(self, request, pb_id, supplier_id):
#         stats = self.get_calculations(pb_id, supplier_id)

#         # Create or update the header
#         negotiation, created = SupplierNegotiation.objects.get_or_create(
#             project_board_id=pb_id,
#             supplier_id=supplier_id
#         )

#         negotiation.gross_total = stats['gross_total']
#         negotiation.net_total = stats['net_total']
#         negotiation.discount_overall_net = stats['discount_overall_net']
#         negotiation.discount_overall_percentage = stats['discount_overall_percentage']
        
#         # Finalize the record
#         negotiation.is_finalized = True
#         negotiation.finalized_by = request.user
#         negotiation.finalized_at = timezone.now()
#         negotiation.finalized_comment = request.data.get('comment', 'Finalized via API')
#         negotiation.save()

#         # Re-use GET to return the full unified structure
#         return self.get(request, pb_id, supplier_id)

#     @transaction.atomic
#     def patch(self, request, pb_id, supplier_id):
#         negotiation = SupplierNegotiation.objects.filter(
#             project_board_id=pb_id, 
#             supplier_id=supplier_id
#         ).first()

#         if not negotiation:
#             return Response({"error": "Negotiation record not found."}, status=status.HTTP_404_NOT_FOUND)

#         # Logic: Reopen negotiation
#         negotiation.is_finalized = False
#         negotiation.finalized_by = None
#         negotiation.finalized_at = None
        
#         reopen_comment = request.data.get('comment', 'Negotiation reopened.')
#         negotiation.finalized_comment = f"Reopened: {reopen_comment}"
#         negotiation.save()

#         # Re-use GET to return the full unified structure
#         return self.get(request, pb_id, supplier_id)


class SupplierNegotiationSummaryAPIView(APIView):
    """
    Handles live calculations, bulk initialization (automatic defaults), 
    finalization, and reopening of supplier negotiations.
    """

    def get_calculations(self, pb_id, supplier_id):
        aggregates = LineItemNegotiation.objects.filter(
            project_board_id=pb_id,
            supplier_id=supplier_id
        ).aggregate(
            sum_gross=Sum('total_price_before'),
            sum_net=Sum('total_price_after'),
            sum_discount_net=Sum('discount_overall_net')
        )

        gross = aggregates['sum_gross'] or Decimal('0.00')
        net = aggregates['sum_net'] or Decimal('0.00')
        discount_net = aggregates['sum_discount_net'] or Decimal('0.00')
        
        percentage = (discount_net / gross * 100) if gross > 0 else Decimal('0.00')

        return {
            "gross_total": gross,
            "net_total": net,
            "discount_overall_net": discount_net,
            "discount_overall_percentage": round(percentage, 2)
        }

    def _get_completion_stats(self, pb_id, supplier_ids=None):
        """
        Calculates expected/agreed counts for specific suppliers or all in a PB.
        """
        total_system_terms = Term.objects.count()
        
        # Base filter
        filters = Q(project_board_id=pb_id)
        if supplier_ids:
            filters &= Q(supplier_id__in=supplier_ids)

        # Aggregate counts
        stats_query = TermAssignment.objects.filter(filters).values('supplier_id').annotate(
            count_expected=Count('term_id', filter=Q(status_type='Expected'), distinct=True),
            count_agreed=Count('term_id', filter=Q(status_type='Agreed'), distinct=True)
        )

        # Map counts to supplier_id
        stats_lookup = {
            entry['supplier_id']: {
                'expected_filled': (entry['count_expected'] >= total_system_terms > 0),
                'agreed_filled': (entry['count_agreed'] >= total_system_terms > 0),
            } for entry in stats_query
        }
        return stats_lookup

    def get(self, request, pb_id, supplier_id=None):
        # 1. Fetch Project Board & Resolve String ID
        try:
            project_board = ProjectBoard.objects.get(id=pb_id)
            commodity_budget = getattr(project_board, 'commodity_budget', 0) or 0
            project_id_str = getattr(project_board, 'project_id_string', str(pb_id)) 
        except ProjectBoard.DoesNotExist:
            return Response({"error": "Project Board not found"}, status=status.HTTP_404_NOT_FOUND)

        # 2. Fetch Rejection History ONCE (Single Query for the whole Board)
        ct = ContentType.objects.get_for_model(SupplierNegotiation)
        rejections = ApprovalRejection.objects.filter(
            pb_id=project_id_str,
            content_type=ct,
            status=ApprovalStatus.REJECTED
        ).order_by('-created_at')
        
        shared_history = ApprovalRejectionSerializer(rejections, many=True).data

        # 3. Fetch Negotiation Headers
        query = SupplierNegotiation.objects.filter(project_board_id=pb_id).select_related('supplier', 'finalized_by')
        
        # --- CASE: Single Supplier Detail ---
        if supplier_id:
            header = query.filter(id=supplier_id).first()
            if not header:
                return Response({"error": "No record found."}, status=status.HTTP_404_NOT_FOUND)
            
            stats = self._get_completion_stats(pb_id, [header.supplier_id])
            result = self._format_from_header(header) # Ensure this method doesn't add history
            
            supplier_stats = stats.get(header.supplier_id, {'expected_filled': False, 'agreed_filled': False})
            result.update(supplier_stats)
            result['commodity_budget'] = commodity_budget
            
            # Keep history here since it's the only result being returned
            result['approval_history'] = shared_history 
            
            return Response(result, status=status.HTTP_200_OK)

        # --- CASE: Bulk Summary ---
        headers = query.all()
        relevant_supplier_ids = [h.supplier_id for h in headers]
        stats_lookup = self._get_completion_stats(pb_id, relevant_supplier_ids)

        results = []
        for h in headers:
            data = self._format_from_header(h)
            data.update(stats_lookup.get(h.supplier_id, {'expected_filled': False, 'agreed_filled': False}))
            # WE REMOVED: data['rejection_history'] = shared_history
            results.append(data)

        # Return history only at the Top Level for the Bulk API
        return Response({
            "project_board_id": pb_id, 
            "commodity_budget": commodity_budget,
            "approval_history": shared_history, 
            "results": results
        }, status=status.HTTP_200_OK)

    def _format_from_header(self, header):
        """ Pulls data ONLY from the SupplierNegotiation fields """
        return {
            "id": header.id,
            "supplier_name": header.supplier.supplier_name,
            "is_finalized": header.is_finalized,
            "totals": {
                "gross": header.gross_total,
                "net": header.net_total,
                "discount": header.discount_overall_net,
                "percentage": header.discount_overall_percentage
            },
            "audit": {
                "at": header.finalized_at,
                "by": header.finalized_by.username if header.finalized_by else None,
                "comment": header.finalized_comment
            }
        }

    @transaction.atomic
    def post(self, request, pb_id, supplier_id=None):
        # 1. Determine which IDs to process
        if supplier_id:
            supplier_ids = [supplier_id]
        else:
            supplier_ids = LineItemNegotiation.objects.filter(
                project_board_id=pb_id
            ).values_list('supplier_id', flat=True).distinct()

        if not supplier_ids:
            return Response({"detail": "No negotiations found."}, status=status.HTTP_404_NOT_FOUND)

        # 2. Unified Processing Logic
        for s_id in supplier_ids:
            stats = self.get_calculations(pb_id, s_id)
            SupplierNegotiation.objects.update_or_create(
                project_board_id=pb_id,
                supplier_id=s_id,
                defaults={
                    'gross_total': stats['gross_total'],
                    'net_total': stats['net_total'],
                    'discount_overall_net': stats['discount_overall_net'],
                    'discount_overall_percentage': stats['discount_overall_percentage'],
                    'is_finalized': False,
                    'finalized_comment': "",
                    'finalized_by': None,
                    'finalized_at': None,
                }
            )

        # 3. Flexible Response
        if supplier_id:
            return self.get(request, pb_id, supplier_id)
        return Response({"message": f"Initialized {len(supplier_ids)} summaries."}, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def patch(self, request, pb_id, supplier_id):
        """ 
        Updates specific fields of the negotiation record, 
        typically used for reopening (unlocking).
        """
        negotiation = SupplierNegotiation.objects.filter(
            project_board_id=pb_id, 
            id=supplier_id
        ).first()

        if not negotiation:
            return Response({"error": "Negotiation record not found."}, status=status.HTTP_404_NOT_FOUND)

        # 1. Get values from the body
        is_finalized_input = request.data.get('is_finalized', False)
        comment_input = request.data.get('comment', "")

        # 2. Apply updates
        negotiation.is_finalized = is_finalized_input
        
        # If we are reopening (setting to False)
        if not is_finalized_input:
            negotiation.finalized_by = None
            negotiation.finalized_at = None
            negotiation.finalized_comment = f"Reopened: {comment_input}"

            # --- RESET ALL APPROVAL REJECTION RECORDS FOR THIS PB ---
            # Removing content_type filter to catch all 15-20 related records
            updated_count = ApprovalRejection.objects.filter(
                pb_id=pb_id
            ).update(
                status=ApprovalStatus.PENDING,
                comments=None,           # Wipe out the comments
                created_at=timezone.now() # Update timestamp to the reset time
            )
            
            print(f"DEBUG: Reset {updated_count} approval records for PB {pb_id}")
            
        else:
            # If for some reason PATCH is used to finalize
            negotiation.finalized_by = request.user
            negotiation.finalized_at = timezone.now()
            negotiation.finalized_comment = comment_input

        negotiation.save()

        # 3. Return the unified structure
        return self.get(request, pb_id, supplier_id)

    # def patch(self, request, pb_id, supplier_id):
    #     """ 
    #     Updates specific fields of the negotiation record, 
    #     typically used for reopening (unlocking).
    #     """
    #     negotiation = SupplierNegotiation.objects.filter(
    #         project_board_id=pb_id, 
    #         id=supplier_id
    #     ).first()

    #     if not negotiation:
    #         return Response({"error": "Negotiation record not found."}, status=status.HTTP_404_NOT_FOUND)

    #     # 1. Get values from the body
    #     is_finalized_input = request.data.get('is_finalized', False)
    #     comment_input = request.data.get('comment', "")

    #     # 2. Apply updates
    #     negotiation.is_finalized = is_finalized_input
        
    #     # If we are reopening (setting to False), clear the metadata
    #     if not is_finalized_input:
    #         negotiation.finalized_by = None
    #         negotiation.finalized_at = None
    #         negotiation.finalized_comment = f"Reopened: {comment_input}"
    #     else:
    #         # If for some reason PATCH is used to finalize
    #         negotiation.finalized_by = request.user
    #         negotiation.finalized_at = timezone.now()
    #         negotiation.finalized_comment = comment_input

    #     negotiation.save()

    #     # 3. Return the unified structure
    #     return self.get(request, pb_id, supplier_id)



class ProcurementUploadAPIView(APIView):
    # MultiPartParser handles form-data (files), JSONParser handles JSON
    parser_classes = (parsers.MultiPartParser, parsers.FormParser)

    def post(self, request, *args, **kwargs):
        serializer = ProcurementUploadSerializer(data=request.data)
        
        if serializer.is_valid():
            # Save while injecting the current user
            serializer.save(uploaded_by=request.user)
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

class ProcurementListByPBView(APIView):
    def get(self, request, pb_id):
        # Filter uploads by the Project Board ID
        uploads = ProcurementUpload.objects.filter(pb_id=pb_id)
        
        # Passing context={'request': request} is the key to getting full URLs
        serializer = ProcurementUploadSerializer(
            uploads, 
            many=True, 
            context={'request': request}
        )
        return Response(serializer.data)

# 2. View to DELETE a specific upload by its own PK
class ProcurementDeleteView(APIView):
    def delete(self, request, pk):
        # Find the specific upload record
        upload = get_object_or_404(ProcurementUpload, pk=pk)
        
        # Optional: Security check to ensure only the owner can delete
        if upload.uploaded_by != request.user:
            return Response(
                {"error": "You do not have permission to delete this file."}, 
                status=status.HTTP_403_FORBIDDEN
            )

        upload.delete()
        return Response(
            {"message": "File deleted successfully."}, 
            status=status.HTTP_204_NO_CONTENT
        )

class TermConditionListCreateAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request):
        # Querying Terms and pre-loading their related conditions
        terms = Term.objects.prefetch_related('conditions').all()
        
        # Serialize the Terms (which now include the condition arrays)
        serializer = TermSerializer(terms, many=True)
        
        return Response(serializer.data)

    def post(self, request):
        serializer = TermConditionSerializer(data=request.data)
        if serializer.is_valid():
            instance = serializer.save(created_by=request.user)
            return Response(TermConditionSerializer(instance).data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)



class TermAssignmentDetailAPIViewAll(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    # def get(self, request, pb_id, supplier_id=None):
    #     # 1. Get status_type from query params (e.g., ?status_type=active)
    #     status_type = request.query_params.get('status_type')

    #     # 2. If supplier_id is missing, find the finalized one for this PB
    #     if not supplier_id:
    #         finalized_negotiation = SupplierNegotiation.objects.filter(
    #             project_board_id=pb_id,
    #             is_finalized=True
    #         ).first()

    #         if finalized_negotiation:
    #             supplier_id = finalized_negotiation.supplier_id
    #         else:
    #             return Response([], status=status.HTTP_200_OK)

    #     # 3. Build the base filter
    #     filters = {
    #         'project_board_id': pb_id,
    #         'supplier_id': supplier_id
    #     }

    #     # 4. Add status_type filter only if it was provided in the URL
    #     if status_type:
    #         filters['status_type'] = status_type

    #     # 5. Fetch records with dynamic filters
    #     line_items = TermAssignment.objects.filter(**filters)
        
    #     serializer = TermAssignmentSummarySerializer(line_items, many=True)
    #     return Response(serializer.data, status=status.HTTP_200_OK)
    def get(self, request, pb_id, supplier_id=None):
        status_type = request.query_params.get('status_type')

        # 1. Resolve supplier_id
        if not supplier_id:
            finalized = SupplierNegotiation.objects.filter(
                project_board_id=pb_id,
                is_finalized=True
            ).first()

            if not finalized:
                return Response([], status=status.HTTP_200_OK)
            supplier_id = finalized.supplier_id

        # 2. Fetch Rejection History as a list
        ct = ContentType.objects.get_for_model(TermAssignment)
        rejections = ApprovalRejection.objects.filter(
            pb_id=pb_id,
            content_type=ct,
            status=ApprovalStatus.REJECTED
        ).order_by('-created_at')

        history_list  = ApprovalRejectionSerializer(rejections, many=True).data


        # 3. Fetch Line Items as a list
        filters = {'project_board_id': pb_id, 'supplier_id': supplier_id}
        if status_type:
            filters['status_type'] = status_type

        line_items = TermAssignment.objects.filter(**filters)
        serializer = TermAssignmentSummarySerializer(line_items, many=True)
        
        # 4. Combine both into one single flat list
        # No keys, just one array of objects
        final_response = list(serializer.data) + history_list

        return Response(final_response, status=status.HTTP_200_OK)

    def patch(self, request, pb_id, supplier_id):
            incoming_data = request.data
            # Ensure it's a list for consistency
            data_list = incoming_data if isinstance(incoming_data, list) else [incoming_data]

            if not data_list:
                return Response({"error": "No data provided."}, status=status.HTTP_400_BAD_REQUEST)

            # Get ContentType for TermAssignment
            ct = ContentType.objects.get_for_model(TermAssignment)

            # Create the Change Request instead of applying it
            # We use a dummy object_id (like pb_id) because lists don't have a single instance ID
            change_req = ChangeRequest.objects.create(
                content_type=ct,
                object_id=pb_id, 
                pending_data=data_list,
                requested_by=request.user,
                is_approved=False
            )

            return Response({
                "message": "Change request submitted for approval.",
                "change_request_id": change_req.id
            }, status=status.HTTP_202_ACCEPTED)
 


class TermAssignmentListCreateView(APIView):
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    def get(self, request):
        # Corrected: Query the actual model intended for the serializer
        assignments = TermAssignment.objects.all()
        serializer = TermAssignmentSummarySerializer(assignments, many=True)
        return Response(serializer.data)


    # def post(self, request):
    #     # 1. Use many=True to handle list input
    #     serializer = TermAssignmentCreateUpdateSerializer(data=request.data, many=True)
        
    #     if not serializer.is_valid():
    #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    #     try:
    #         with transaction.atomic():
    #             # 2. Save all TermAssignment instances (bulk create)
    #             # Inject created_by for each
    #             instances = []
    #             for item_data in serializer.validated_data:
    #                 item = TermAssignment(**item_data)
    #                 item.created_by = request.user
    #                 item.save()
    #                 instances.append(item)

    #             # 3. For each created TermAssignment, create an approved ChangeRequest
    #             content_type = ContentType.objects.get_for_model(TermAssignment)
                
    #             change_requests = []
    #             # Match raw input data with created instances (assuming input order matches)
    #             raw_data_list = request.data if isinstance(request.data, list) else [request.data]
                
    #             for idx, instance in enumerate(instances):
    #                 raw_data = raw_data_list[idx] if idx < len(raw_data_list) else {}
                    
    #                 cr = ChangeRequest.objects.create(
    #                     content_type=content_type,
    #                     object_id=instance.id,
    #                     pending_data=raw_data,  # raw data for this specific row
    #                     requested_by=request.user,
    #                     is_approved=True,       # creation is auto-approved
    #                 )
    #                 change_requests.append(cr.id)

    #             # 4. Return the created records (summary serializer)
    #             return Response(
    #                 {
    #                     "message": f"{len(instances)} TermAssignment(s) created successfully",
    #                     "term_assignment_ids": [i.id for i in instances],
    #                     "change_request_ids": change_requests
    #                 },
    #                 status=status.HTTP_201_CREATED
    #             )

    #     except Exception as e:
    #         return Response(
    #             {"error": "Failed to create TermAssignment(s)", "detail": str(e)},
    #             status=status.HTTP_500_INTERNAL_SERVER_ERROR
    #         )
    def post(self, request):
            serializer = TermAssignmentCreateUpdateSerializer(data=request.data, many=True)
            if not serializer.is_valid():
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

            # Helper to handle Decimal/Date serialization
            def serialize_data(data):
                return json.loads(json.dumps(data, cls=DjangoJSONEncoder))

            try:
                with transaction.atomic():
                    # 1. Bulk Create TermAssignments
                    instances = []
                    for item_data in serializer.validated_data:
                        item = TermAssignment(**item_data)
                        item.created_by = request.user
                        item.save()
                        instances.append(item)

                    # 2. Create ONE SINGLE ChangeRequest for the entire batch
                    if instances:
                        content_type = ContentType.objects.get_for_model(TermAssignment)
                        
                        # Store the entire request.data (the whole list) in one record
                    ChangeRequest.objects.create(
                        content_type=content_type,
                        # We anchor to the first instance id to satisfy the field
                        object_id=instances[0].id, 
                        pending_data=serialize_data(request.data), 
                        requested_by=request.user,
                        is_approved=True,
                    )

                    return Response(
                        {
                            "message": f"{len(instances)} TermAssignment(s) created with a single audit record.",
                            "term_assignment_ids": [i.id for i in instances],
                        },
                        status=status.HTTP_201_CREATED
                    )

            except Exception as e:
                return Response(
                    {"error": "Failed to create TermAssignment(s)", "detail": str(e)},
                    status=status.HTTP_500_INTERNAL_SERVER_ERROR
                )



class TermAssignmentDetailAPIView(APIView): # Renamed for clarity
    def get_permissions(self):
        if self.request.method == 'GET':
            return [AllowAny()]
        return super().get_permissions()

    def get(self, request, id):
        # Changed LineItem to TermAssignment to match your serializers
        assignment = get_object_or_404(TermAssignment, id=id)
        serializer = TermAssignmentSummarySerializer(assignment)
        return Response(serializer.data, status=status.HTTP_200_OK)

    # def patch(self, request, id=None):
    #     # 1. Capture and Normalize data
    #     incoming_data = request.data
    #     is_list = isinstance(incoming_data, list)
    #     data_to_process = incoming_data if is_list else [incoming_data]

    #     # 2. Global Validation
    #     # Check if the list structure is correct before we start touching the DB
    #     serializer = TermAssignmentCreateUpdateSerializer(
    #         data=data_to_process, 
    #         many=True, 
    #         partial=True
    #     )
        
    #     if not serializer.is_valid():
    #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    #     updated_info = []

    #     try:
    #         # 3. Loop and Update
    #         for item_data in data_to_process:
    #             # Priority: ID from the JSON body, fallback to ID from the URL
    #             assignment_id = item_data.get('id') or id
                
    #             if not assignment_id:
    #                 raise Exception("Object ID is required (either in the body or the URL).")

    #             # Fetch the actual record
    #             assignment = get_object_or_404(TermAssignment, id=assignment_id)

    #             # 4. Apply changes using a single-instance serializer
    #             # This handles the "Integer vs Instance" issue for ForeignKeys automatically
    #             instance_serializer = TermAssignmentCreateUpdateSerializer(
    #                 assignment, 
    #                 data=item_data, 
    #                 partial=True
    #             )
                
    #             if instance_serializer.is_valid():
    #                 instance_serializer.save()
    #             else:
    #                 # If this specific instance fails, raise error to trigger rollback
    #                 raise Exception(instance_serializer.errors)
                
    #             updated_info.append({
    #                 "id": assignment.id, 
    #                 "status": "Updated successfully"
    #             })

    #         # 5. Final Response
    #         return Response({
    #             "message": f"Successfully updated {len(updated_info)} records.",
    #             "details": updated_info if is_list else updated_info[0]
    #         }, status=status.HTTP_200_OK)

    #     except Exception as e:
    #         # transaction.atomic ensures if one fails, none are saved
    #         return Response(
    #             {"error": str(e)}, 
    #             status=status.HTTP_400_BAD_REQUEST
    #         )
    def patch(self, request, id=None):
        incoming_data = request.data
        is_list = isinstance(incoming_data, list)
        data_to_process = incoming_data if is_list else [incoming_data]

        serializer = TermAssignmentCreateUpdateSerializer(
            data=data_to_process, 
            many=True, 
            partial=True
        )
        
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        updated_info = []
        ct = ContentType.objects.get_for_model(TermAssignment)

        try:
            with transaction.atomic():
                # 1. PRE-CHECK: Is the board locked?
                # We peek at the first item to get the Project Board ID
                first_item_id = data_to_process[0].get('id') or id
                first_assignment = get_object_or_404(TermAssignment, id=first_item_id)
                target_pb_id = first_assignment.project_board_id

                existing_request = ChangeRequest.objects.filter(
                    content_type=ct,
                    object_id=target_pb_id, # PB is the anchor
                    is_approved=False
                ).exists()

                if existing_request:
                    return Response(
                        {"error": "This board is locked due to a pending change request."}, 
                        status=status.HTTP_400_BAD_REQUEST
                    )

                # 2. PROCESS ITEMS
                for item_data in data_to_process:
                    assignment_id = item_data.get('id') or id
                    assignment = get_object_or_404(TermAssignment, id=assignment_id)
                    
                    instance_serializer = TermAssignmentCreateUpdateSerializer(
                        assignment, 
                        data=item_data, 
                        partial=True
                    )
                    
                    if instance_serializer.is_valid():
                        instance_serializer.save()
                        
                        # Update THIS specific section to Pending if it was rejected
                        ApprovalRejection.objects.filter(
                            pb_id=target_pb_id, 
                            content_type=ct,
                            status=ApprovalStatus.REJECTED
                        ).update(status=ApprovalStatus.PENDING)
                    else:
                        raise Exception(instance_serializer.errors)
                    
                    updated_info.append({"id": assignment.id, "status": "Updated"})

                # 3. FINAL GLOBAL CHECK
                # If NO rejections remain for this PB, wipe all comments/reset all to Pending
                other_rejected_exists = ApprovalRejection.objects.filter(
                    pb_id=target_pb_id,
                    status=ApprovalStatus.REJECTED
                ).exists()

                if not other_rejected_exists:
                    ApprovalRejection.objects.filter(pb_id=target_pb_id).update(
                        status=ApprovalStatus.PENDING,
                        comments=None,
                        created_at=timezone.now()
                    )

            return Response({
                "message": f"Successfully updated {len(updated_info)} records.",
                "details": updated_info if is_list else updated_info[0]
            }, status=status.HTTP_200_OK)

        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
 
class ApprovarListCreateAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request, project_id):
        pbs = Approver.objects.filter(project_board_id=project_id)
        return Response(ApproverSummarySerializer(pbs, many=True).data)

    # def post(self, request):
    #     serializer = ApproverCreateUpdateSerializer(data=request.data)
        
    #     if not serializer.is_valid():
    #         return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    #     try:
    #         with transaction.atomic():
    #             # 1. Create the Approver record
    #             approver = serializer.save()

    #             # 2. Automatically create an approved ChangeRequest (audit/history)
    #             content_type = ContentType.objects.get_for_model(Approver)
                

    #             # 2. Use request.user ONLY for the ChangeRequest
    #             ChangeRequest.objects.create(
    #                 content_type=ContentType.objects.get_for_model(Approver),
    #                 object_id=approver.id,
    #                 pending_data=request.data.copy(),
    #                 requested_by=request.user,  # It still gets recorded here!
    #                 is_approved=True,
    #             )

    #             # 3. Return the created approver
    #             return Response(
    #                 ApproverSummarySerializer(approver).data,
    #                 status=status.HTTP_201_CREATED
    #             )

    #     except Exception as e:
    #         return Response(
    #             {"error": "Failed to create Approver", "detail": str(e)},
    #             status=status.HTTP_500_INTERNAL_SERVER_ERROR
    #         )
    def post(self, request):
        is_automatic = request.data.get('is_automatic', False)
        pb_id = request.data.get('project_board')
        
        if not pb_id:
            return Response({"error": "project_board ID is required"}, status=status.HTTP_400_BAD_REQUEST)

        pb = get_object_or_404(ProjectBoard, id=pb_id)

        try:
            with transaction.atomic():
                # --- STRATEGY A: AUTOMATIC HIERARCHY ---
                if is_automatic:
                    print(f"\n--- Starting Automatic Hierarchy for PB: {pb_id} ---")
                    
                    # 1. Clear existing approvers
                    deleted_count = Approver.objects.filter(project_board=pb).delete()
                    print(f"Cleared {deleted_count[0]} existing approvers.")

                    # 2. Calculate Total Value
                    # We filter finalized negotiations and sum their net_total
                    aggregates = SupplierNegotiation.objects.filter(
                        project_board=pb,
                        is_finalized=True
                    ).aggregate(total=Sum('net_total'))
                    
                    total_val = aggregates['total'] or 0
                    print(f"DEBUG: Calculated Total Value = {total_val}")

                    # 3. Define the Hierarchy structure
                    hierarchy_data = []
                    
                    # Level 1: Project Manager
                    if pb.project and pb.project.project_manager:
                        pm_username = pb.project.project_manager.username
                        hierarchy_data.append((pm_username, 1, "MANDATORY"))
                        print(f"Level 1 added: PM ({pm_username})")
                    else:
                        print("Level 1 skipped: No Project Manager assigned to project.")
                    
                    # Level 2: Santosh
                    hierarchy_data.append(('sbadam', 2, "MANDATORY"))
                    print("Level 2 added: sbadam (Default)")
                    
                    # Level 3: Niranjan (If > 2L)
                    if total_val > 200000:
                        hierarchy_data.append(('npanigrahi', 3, "MANDATORY"))
                        print(f"Level 3 added: npanigrahi (Total {total_val} > 200,000)")
                    else:
                        print(f"Level 3 skipped: Total {total_val} <= 200,000")

                    # Level 4: Raghuraj (If >= 10L)
                    if total_val >= 1000000:
                        hierarchy_data.append(('rdeshpande', 4, "MANDATORY"))
                        print(f"Level 4 added: rdeshpande (Total {total_val} >= 1,000,000)")
                    else:
                        print(f"Level 4 skipped: Total {total_val} < 1,000,000")

                    # 4. Create Approver objects
                    created_approvers = []
                    for username, level, app_type in hierarchy_data:
                        user = User.objects.filter(username=username).first()
                        if user:
                            app_obj = Approver.objects.create(
                                project_board=pb,
                                approver=user,
                                level=level,
                                email=user.email,
                                approval_type=app_type
                            )
                            created_approvers.append(ApproverSummarySerializer(app_obj).data)
                            print(f"Successfully created Approver record for: {username} at Level {level}")
                        else:
                            print(f"WARNING: User '{username}' not found in database. Skipping level {level}.")

                    print(f"--- Hierarchy Complete. Total Approvers: {len(created_approvers)} ---\n")

                    return Response({
                        "message": f"Hierarchy applied for total: {total_val}",
                        "total_value_calculated": total_val,
                        "approvers": created_approvers
                    }, status=status.HTTP_200_OK)

                # --- STRATEGY B: SINGLE MANUAL POST ---
                else:
                    print("DEBUG: Manual Strategy Selected")
                    serializer = ApproverCreateUpdateSerializer(data=request.data)
                    if serializer.is_valid():
                        approver = serializer.save()
                        
                        ChangeRequest.objects.create(
                            content_type=ContentType.objects.get_for_model(Approver),
                            object_id=approver.id,
                            pending_data=request.data.copy(),
                            requested_by=request.user,
                            is_approved=True,
                        )
                        return Response(ApproverSummarySerializer(approver).data, status=status.HTTP_201_CREATED)
                    
                    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            print(f"ERROR in Approver Post: {str(e)}")
            import traceback
            traceback.print_exc() # This will print the full error stack trace to your console
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

class ApprovarDetailAPIView(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request, pb_id):
        # Changed LineItem to TermAssignment to match your serializers
        approvers = Approver.objects.filter(project_board=pb_id)
        serializer = ApproverSummarySerializer(approvers,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


class ApprovarDetailAPIViewId(APIView):
    def get_permissions(self):
        return [permissions.AllowAny()] if self.request.method == 'GET' else [permissions.IsAuthenticated()]

    def get(self, request, id):
        # Changed LineItem to TermAssignment to match your serializers
        approvers = Approver.objects.filter(id=id)
        serializer = ApproverSummarySerializer(approvers,many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def patch(self, request, id):
        pb = get_object_or_404(Approver, id=id)
        
        # 1. Validate data against the model's rules
        serializer = ApproverCreateUpdateSerializer(pb, data=request.data, partial=True)
        
        if serializer.is_valid():
            # 2. Get ContentType for the generic foreign key
            content_type = ContentType.objects.get_for_model(pb)
            
            existing_request = ChangeRequest.objects.filter(
                content_type=content_type,
                object_id=pb.id,
                is_approved=False
            ).exists()

            if existing_request:
                return Response(
                    {"error": "A pending change request already exists for this record. Please wait for approval or rejection."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            change_request = ChangeRequest.objects.create(
                content_type=content_type,
                object_id=pb.id,
                pending_data=request.data, # Store the raw updated fields
                requested_by=request.user,
                is_approved=False
            )
            
            return Response({
                "message": "Update intercepted. Change request created.",
                "change_request_id": change_request.id,
                "status": "Pending Approval"
            }, status=status.HTTP_202_ACCEPTED)
            
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # def patch(self, request, pb_id):
    #     pb = get_object_or_404(ProjectBoard, id=pb_id)
    #     serializer = ProjectBoardCreateUpdateSerializer(pb, data=request.data, partial=True)
    #     if serializer.is_valid():
    #         instance = serializer.save()
    #         return Response(ProjectBoardReadSerializer(instance).data)
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)




class ProjectBoardSummaryAPIView(APIView):
    def get(self, request, pb_id):
        # Optimized with select_related to reduce DB queries
        pb = get_object_or_404(
            ProjectBoard.objects.select_related('project', 'buyer'), 
            id=pb_id
        )
        serializer = PBSummarySerializer(pb)
        return Response(serializer.data)


class GenerateApprovalMatrixView(APIView):
    @transaction.atomic
    def post(self, request, pb_id):
        pb = get_object_or_404(ProjectBoard.objects.select_related("project", "buyer"), id=pb_id)
        negotiation = SupplierNegotiation.objects.filter(project_board=pb, is_finalized=True).first()
        
        if not negotiation:
            return Response({"message": "No finalized supplier found"}, status=400)

        # Matrix Logic
        Approver.objects.filter(project_board=pb).delete()
        # Define roles (Santosh logic helper used here)
        santosh = User.objects.filter(first_name__iexact="Santosh", role="project_manager_hod").first()
        
        flow = [(1, pb.buyer, "MANDATORY"), (2, santosh, "MANDATORY"), (3, pb.project.project_manager, "MANDATORY")]
        
        for level, user, app_type in flow:
            if user:
                Approver.objects.create(project_board=pb, approver=user, email=user.email, level=level, approval_type=app_type)

        pb.pb_status = "APPROVAL"
        pb.save()
        return Response({"message": "Matrix generated"}, status=201)

class MyApprovalListView(APIView):
    def get(self, request, user_id):
        approvals = Approver.objects.filter(approver_id=user_id, project_board__pb_status="APPROVAL").select_related('project_board__project')
        result = []
        for app in approvals:
            # Logic: Can only see if previous level is Approved
            if app.level > 1:
                prev_approved = Approver.objects.filter(project_board=app.project_board, level=app.level-1, status="Approved").exists()
                if not prev_approved: continue
            
            result.append({
                "project_name": app.project_board.project.project.project_name if app.project_board.project else None,
                "pb_unique_code": app.project_board.document_no,
                "approval_level": app.level,
                "status": app.status
            })
        return Response({"count": len(result), "data": result})

class TakeApprovalActionView(APIView):
    @transaction.atomic
    def post(self, request):
        pb_id = request.data.get("pb_id")
        approver_id = request.data.get("approver_id")
        action = request.data.get("action") # APPROVED / REJECTED

        approver = get_object_or_404(Approver, project_board_id=pb_id, approver_id=approver_id)
        
        # Logic: Sequential check
        if Approver.objects.filter(project_board_id=pb_id, level__lt=approver.level, status="Pending").exists():
            return Response({"message": "Previous level pending"}, status=400)

        approver.status = "Approved" if action == "APPROVED" else "Rejected"
        approver.comments = request.data.get("comment")
        approver.action_taken_at = timezone.now()
        approver.save()

        # Update Board Status
        pb = approver.project_board
        if action == "REJECTED":
            pb.pb_status = "REJECTED"
        else:
            pending = Approver.objects.filter(project_board=pb, status="Pending").exists()
            pb.pb_status = "APPROVED" if not pending else "APPROVAL"
        pb.save()

        return Response({"message": "Action recorded", "pb_status": pb.pb_status})

class ProjectBoardFullDetailsView(APIView):
    def get(self, request, pb_id):
        pb = get_object_or_404(ProjectBoard.objects.select_related("project"), id=pb_id)
        
        # Prefetching and organizing data
        response_data = {
            "basic_details": ProjectBoardBasicSerializer(pb).data,
            "line_items": ProjectBoardLineItemSerializer(LineItem.objects.filter(project_board=pb), many=True).data,
            "approvals": ApproverBaseSerializer(Approver.objects.filter(project_board=pb).order_by('level'), many=True).data,
            "supplier_finalized": SupplierFinalizedSerializer(SupplierNegotiation.objects.filter(project_board=pb, is_finalized=True), many=True).data
        }
        return Response(response_data)

class DraftListView(APIView):

    def post(self, request):
        table_name = request.data.get('table_name', '').lower()
        # For new records, frontend should send object_id as 0 or null
        object_id = request.data.get('object_id')
        pending_data = request.data.get('data')
        draft_id = request.data.get('draft_id')
        # 1. Validate Table Name
        try:
            ct = ContentType.objects.get(app_label='purchase', model=table_name)            
            allowed_models = [
                'projectboard', 'commercialrequirement', 'supplier', 
                'suppliernegotiation', 'lineitem', 'lineitemnegotiation',
                'term', 'termassignment','termcondition','approver'
            ]
            if table_name not in allowed_models:
                return Response({"error": "Table not allowed."}, status=status.HTTP_400_BAD_REQUEST)
        except ContentType.DoesNotExist:
            return Response({"error": "Invalid table name."}, status=status.HTTP_400_BAD_REQUEST)

        # 2. Conditional Verification
        # If object_id is provided and > 0, we check if it exists (Update mode)
        # If object_id is 0 or None, we skip this (Create mode)
        if object_id and int(object_id) > 0:
            model_class = ct.model_class()
            # if not model_class.objects.filter(id=object_id).exists():
            #     return Response({"error": f"Record with ID {object_id} not found in {table_name}."}, status=status.HTTP_404_NOT_FOUND)
        else:
            # Set object_id to 0 to represent a "New Record Draft"
            object_id = 0

        if draft_id:
            # We already have a draft, so we just update it
            try:
                draft = Draft.objects.get(id=draft_id, requested_by=request.user)
                draft.pending_data = pending_data
                draft.save()
                created = False
            except Draft.DoesNotExist:
                return Response({"error": "Draft not found"}, status=status.HTTP_404_NOT_FOUND)
        else:
            # First time saving - create the draft row
            draft = Draft.objects.create(
                content_type=ct,
                object_id=object_id or 0,
                is_approved=False,
                requested_by=request.user,
                pending_data=pending_data
            )
            created = True

        # 3. Response: NOW we send the ID back to the frontend
        return Response({
            "message": "Draft created" if created else "Draft updated",
            "draft_id": draft.id,
            "mode": "create" if draft.object_id == 0 else "update"
            }, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)

class DraftDetailView(APIView):
    """
    Retrieve a specific draft by its Primary Key (ID) via GET.
    URL Pattern: path('draft/<int:pk>/', DraftDetailView.as_view())
    """
    def get(self, request, pk):
        # 1. Fetch the draft by ID
        # We filter by requested_by to ensure users can only see their own drafts
        draft = get_object_or_404(Draft, id=pk, requested_by=request.user)

        # 2. Determine if it's a draft for a new record or an update
        # Since you use 0 to represent a "New Record Draft"
        mode = "create" if str(draft.object_id) == "0" else "update"
        
        # 3. Return the response
        return Response({
            "draft_id": draft.id,
            "table_name": draft.content_type.model,
            "object_id": draft.object_id,
            "data": draft.pending_data,
            "mode": mode,
            "requested_by": draft.requested_by.username,
            "created_at": draft.created_at,
        }, status=status.HTTP_200_OK)

    def delete(self, request, pk):
            # 1. Fetch the draft (ensuring it belongs to the user)
            draft = get_object_or_404(Draft, id=pk, requested_by=request.user)
            
            # 2. Perform the deletion
            draft_id = draft.id  # Store ID for the response message
            draft.delete()
            
            # 3. Return a success message
            return Response({
                "message": f"Draft {draft_id} has been deleted successfully.",
                "status": "success"
            }, status=status.HTTP_204_NO_CONTENT)

class ChangeRequestListView(APIView):
    def get(self, request):
        table_name = request.query_params.get('table_name')
        target_id = request.query_params.get('pb_id') or request.query_params.get('id')
        
        queryset = ChangeRequest.objects.filter(is_approved=False)

        # 1. Table Filtering
        if table_name:
            queryset = queryset.filter(content_type__model=table_name.lower())


        if target_id:
            from django.db.models import Q
            
            # Filter A: The direct object_id column
            query = Q(object_id=target_id)
            
            # Filter B: If pending_data is a Dictionary
            query |= Q(pending_data__project_board=int(target_id))
            query |= Q(pending_data__project_board=str(target_id))

            # Filter C: If pending_data is a LIST (For Bulk TermAssignments)
            # This checks if the list contains an object with this project_board ID
            query |= Q(pending_data__contains=[{'project_board': int(target_id)}])
            query |= Q(pending_data__contains=[{'project_board': str(target_id)}])

            queryset = queryset.filter(query)
        # 2. Smart ID Filtering
        # if target_id:
        #     from django.db.models import Q
            
        #     # Use Q objects to check BOTH:
        #     # A) The object_id column (for updates to existing records)
        #     # B) The JSON pending_data field (for new record drafts)
        #     queryset = queryset.filter(
        #         Q(object_id=str(target_id)) | 
        #         Q(pending_data__project_board=int(target_id)) |
        #         Q(pending_data__project_board=str(target_id))
        #     )

        # 3. Build Response
        # results = []
        # for cr in queryset:
        #     results.append({
        #         "id": cr.id, 
        #         "table_name": cr.content_type.model,
        #         "object_id": cr.object_id,
        #         "pending_data": cr.pending_data,
        #         "requested_by": cr.requested_by.username if cr.requested_by else None,
        #         "created_at": cr.created_at.strftime("%Y-%m-%d %H:%M:%S")
        #     })

        # return Response(results, status=status.HTTP_200_OK)
        results = []
        for cr in queryset:
            model_key = cr.content_type.model.lower()
            raw_json = cr.pending_data
            
            # --- HYDRATION LOGIC ---
            hydrator_class = HYDRATION_MAP.get(model_key)
            display_data = raw_json # Fallback

            if hydrator_class and raw_json:
                is_bulk = isinstance(raw_json, list)
                
                # We initialize with data=raw_json
                # Since we are using these purely for representation:
                instance_hydrator = hydrator_class(data=raw_json, many=is_bulk)
                
                try:
                    # We use to_representation to force ID -> Name lookup
                    if is_bulk:
                        display_data = [instance_hydrator.child.to_representation(item) for item in raw_json]
                    else:
                        display_data = instance_hydrator.to_representation(raw_json)
                except Exception:
                    display_data = raw_json # If ID doesn't exist in DB, show raw ID

            results.append({
                "id": cr.id, 
                "table_name": model_key,
                "object_id": cr.object_id,
                "pending_data": display_data, # Now contains names
                "requested_by": cr.requested_by.username if cr.requested_by else None,
                "created_at": cr.created_at.strftime("%Y-%m-%d %H:%M:%S")
            })

        return Response(results, status=status.HTTP_200_OK)

@csrf_exempt
def get_single_project_data(request, project_id):
    # 1. Fetch the main Project
    project_instance = get_object_or_404(Project, id=project_id)
    project_name = project_instance.project.project_name if project_instance.project else f"Project {project_instance.id}"

    # --- 2. Process Cash Inflows (Project Level) ---
    inflows_qs = CashInflow.objects.filter(project=project_instance).annotate(
        year=ExtractYear('date'),
        month=ExtractMonth('date')
    ).values('year', 'month').annotate(total=Sum('amount'))

    inflow_map = defaultdict(float)
    for entry in inflows_qs:
        inflow_map[(entry['year'], entry['month'])] = float(entry['total'])

    # --- 3. Process Procurements (ProjectBoard Level) ---
    procurements_data = []
    boards = project_instance.project_boards.all() 
    project_total_outflow = defaultdict(float)

    for board in boards:
        # 1. Fetch the finalized Supplier Negotiation for this board
        finalized_negotiation = board.supplier_negotiations.filter(is_finalized=True).first()
        
        # If no finalized negotiation, budget is 0. 
        # (Optional: Use .net_total or .gross_total based on your accounting preference)
        supplier_budget = float(finalized_negotiation.net_total) if finalized_negotiation else 0.0

        # 2. Fetch Payment Term percentages tied to this board
        # Ensure TermAssignment has a date and a percentage field
        outflows_qs = TermAssignment.objects.filter(
            project_board=board,
            term__name__icontains="Payment",
            status_type="Agreed"  # <--- Add this filter to exclude 'Expected'
        ).annotate(
            year=ExtractYear('date'),
            month=ExtractMonth('date')
        ).values('year', 'month').annotate(total_pct=Sum('percentage'))
        board_monthly = []
        for out in outflows_qs:
            yr, mn = out['year'], out['month']
            
            # 3. CALCULATE OUTFLOW: (Budget * Percentage / 100)
            # Use Decimal for precision if possible, but keeping float as per your snippet
            percentage_value = float(out['total_pct'])
            calculated_amount = supplier_budget * (percentage_value / 100)
            
            board_monthly.append({
                "year": yr,
                "month": mn,
                "percentage": percentage_value,
                "outflow_amount": round(calculated_amount, 2)
            })
            
            # Aggregate to the project-wide total
            project_total_outflow[(yr, mn)] += calculated_amount

        procurements_data.append({
            "procurementId": board.id,
            "documentNo": board.document_no,
            "commodity": board.commodity_equipment,
            "supplierBudget": supplier_budget,
            "monthlyOutflows": board_monthly
        })

    # --- 4. Construct Final Summary ---
    all_periods = sorted(set(inflow_map.keys()) | set(project_total_outflow.keys()))
    
    total_summary = []
    for yr, mn in all_periods:
        total_summary.append({
            "year": yr,
            "month": mn,
            "totalInflow": inflow_map[(yr, mn)],
            "totalOutflow": project_total_outflow[(yr, mn)]
        })

    response = {
        "projectId": project_instance.id,
        "projectName": project_name,
        "summaryByYearMonth": total_summary,
        "procurements": procurements_data
    }

    return JsonResponse(response)


class ApprovalRejectionAPIView(APIView):



    # def get(self, request, pk=None):
    #     print(f"DEBUG: Logged in User: {request.user}")
    #     print(f"DEBUG: PK from URL: {pk}")
    #     print(f"DEBUG: pb_id from Query Params: {request.query_params.get('pb_id')}")

    #     if pk:
    #         # Try to get it without the user filter first to see if it exists at all
    #         instance = get_object_or_404(ApprovalRejection, pk=pk)
    #         print(f"DEBUG: Found instance. Created by: {instance.created_by}")
    #         serializer = ApprovalRejectionSerializer(instance)
    #         return Response(serializer.data)

    #     # Listing
    #     queryset = ApprovalRejection.objects.all() # Start with ALL
    #     print(f"DEBUG: Total records in DB: {queryset.count()}")
        
    #     # Then filter step by step
    #     user_queryset = queryset.filter(created_by=request.user)
    #     print(f"DEBUG: Records for this user: {user_queryset.count()}")

    #     pb_id = request.query_params.get('pb_id')
    #     if pb_id:
    #         user_queryset = user_queryset.filter(pb_id=pb_id)
    #         print(f"DEBUG: Records for user + pb_id {pb_id}: {user_queryset.count()}")

    #     serializer = ApprovalRejectionSerializer(user_queryset, many=True)
    #     return Response(serializer.data)
    def get(self, request, pk=None):
        # 1. Check for "Procurement Admin" status
        is_procurement_admin = UserModulePermission.objects.filter(
            user=request.user,
            module__code='Procurement',  # Assuming 'procurement' is the code in your Module model
            permission__code='Admin',    # Assuming 'admin' is the code in your ModulePermission model
            is_active=True
        ).exists()

        # Define if the user should behave like a Superuser
        is_privileged_user = request.user.is_superuser or is_procurement_admin

        print(f"DEBUG: User: {request.user} | Superuser: {request.user.is_superuser} | ProcAdmin: {is_procurement_admin}")

        if pk:
            instance = get_object_or_404(ApprovalRejection, pk=pk)
            
            # Check ownership unless the user is privileged
            if not is_privileged_user and instance.created_by != request.user:
                return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)
            
            serializer = ApprovalRejectionSerializer(instance)
            return Response(serializer.data)

        # 2. Start with the base queryset
        queryset = ApprovalRejection.objects.all()

        # 3. ROLE CHECK: If NOT a privileged user, filter by created_by
        if not is_privileged_user:
            queryset = queryset.filter(created_by=request.user)
            print(f"DEBUG: Filtering applied. Count: {queryset.count()}")
        else:
            print("DEBUG: Privileged user detected. Accessing all records.")

        # 4. Filter by pb_id if provided
        pb_id = request.query_params.get('pb_id')
        if pb_id:
            queryset = queryset.filter(pb_id=pb_id)

        serializer = ApprovalRejectionSerializer(queryset, many=True)
        return Response(serializer.data)


    def post(self, request):
        """Create a new approval request."""
        data = request.data
        model_name = data.get('model_name')
        
        try:
            ctype = ContentType.objects.get(model=model_name.lower())
        except (ContentType.DoesNotExist, AttributeError):
            return Response(
                {"error": f"Invalid model name: {model_name}"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        serializer = ApprovalRejectionSerializer(data=data)
        if serializer.is_valid():
            serializer.save(
                content_type=ctype,
                created_by=request.user if request.user.is_authenticated else None
            )
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    # def patch(self, request, pk=None):
    #     """Update a specific record via URL ID."""
    #     if not pk:
    #         return Response({"error": "ID required in URL"}, status=status.HTTP_400_BAD_REQUEST)

    #     instance = get_object_or_404(ApprovalRejection, pk=pk)
        
    #     # partial=True allows you to update only the 'status' or 'comment'
    #     serializer = ApprovalRejectionSerializer(instance, data=request.data, partial=True)
    #     if serializer.is_valid():
    #         serializer.save()
    #         return Response(serializer.data)
        
    #     return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def patch(self, request, pk=None):
        """
        Update logic:
        - Superuser: Can ONLY update 'status'.
        - Regular User: Can update any field (standard partial update).
        """
        if not pk:
            return Response({"error": "ID required in URL"}, status=status.HTTP_400_BAD_REQUEST)

        instance = get_object_or_404(ApprovalRejection, pk=pk)

        # 1. Prepare the data based on the User Role
        if request.user.is_superuser:
            # If they are a superuser, we strip everything except 'status'
            status_value = request.data.get('status')
            if status_value is None:
                return Response(
                    {"error": "Superusers are only permitted to update the 'status' field."}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
            patch_data = {'status': status_value}
        else:
            # Regular user gets to update whatever they sent in the request
            patch_data = request.data

        # 2. Apply the update using the filtered (or unfiltered) data
        serializer = ApprovalRejectionSerializer(instance, data=patch_data, partial=True)
        
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


class BulkInitializeWorkflowAPIView(APIView):
    def post(self, request):
        pb_id = request.data.get('pb_id')
        
        if not pb_id:
            return Response({"error": "pb_id is required"}, status=400)

        try:
            with transaction.atomic():
                # 1. Get the Project Board
                pb = ProjectBoard.objects.get(id=pb_id)

                # 2. Get the 5 specific ContentTypes
                model_list = [
                    'projectboard', 'commercialrequirement', 
                    'suppliernegotiation', 'lineitemnegotiation', 'termassignment'
                ]
                content_types = ContentType.objects.filter(model__in=model_list)

                # 3. Get all Approvers for this PB
                approvers = Approver.objects.filter(project_board=pb)

                if not approvers.exists():
                    return Response({"error": "No approvers found for this PB"}, status=400)

                created_count = 0
                
                # 4. Nested Loop: For each Approver, create a record for each Model
                for approver_record in approvers:
                    user = approver_record.approver
                    
                    for ct in content_types:
                        # get_or_create prevents duplicates if API is hit twice
                        obj, created = ApprovalRejection.objects.get_or_create(
                            pb_id=pb,
                            content_type=ct,
                            created_by=user,
                            defaults={
                                'status': ApprovalStatus.PENDING,
                                'comments': '' # Blank as requested
                            }
                        )
                        if created:
                            created_count += 1

            return Response({
                "message": f"Initialization complete. Created {created_count} approval records.",
                "total_users": approvers.count(),
                "total_models": content_types.count()
            }, status=status.HTTP_201_CREATED)

        except ProjectBoard.DoesNotExist:
            return Response({"error": "Project Board not found"}, status=404)
        except Exception as e:
            return Response({"error": str(e)}, status=500)        

# class ChangeRequestPatchView(generics.UpdateAPIView):
#     queryset = ChangeRequest.objects.all()
#     serializer_class = ChangeRequestApprovalSerializer

#     def update(self, request, *args, **kwargs):
#         instance = self.get_object()
        
#         if instance.is_approved:
#             return Response({"error": "Already approved."}, status=400)

#         is_approving = request.data.get('is_approved', False)

#         try:
#             with transaction.atomic():
#                 # 1. Update the ChangeRequest status
#                 serializer = self.get_serializer(instance, data=request.data, partial=True)
#                 serializer.is_valid(raise_exception=True)
#                 cr_instance = serializer.save() 

#                 # 2. If approved, sync the JSON data back to the real Models
#                 if is_approving:
#                     target_obj = cr_instance.content_object 
#                     pending_data = cr_instance.pending_data

#                     # --- CASE: LINE ITEM ---
#                     if isinstance(target_obj, LineItem):
#                         # Update LineItem core fields
#                         for field, value in pending_data.items():
#                             if field == 'negotiations':
#                                 continue
                                
#                             # Check if the field is a ForeignKey by looking for it in the model's fields
#                             # If it is a ForeignKey, append '_id' to the attribute name
#                             field_obj = target_obj._meta.get_field(field)
#                             if field_obj.is_relation and not field_obj.many_to_many:
#                                 target_attr = f"{field}_id"
#                             else:
#                                 target_attr = field

#                             if hasattr(target_obj, target_attr):
#                                 setattr(target_obj, target_attr, value)
                        
#                         # Save LineItem first
#                         target_obj.save()

#                         # Update related Negotiations
#                         negotiations_list = pending_data.get('negotiations', [])
#                         for neg_data in negotiations_list:
#                             neg_id = neg_data.get('id')
#                             if neg_id:
#                                 # Fetch the actual Negotiation record
#                                 # select_related('line_item') ensures calculate_totals has the data it needs
#                                 neg_instance = LineItemNegotiation.objects.select_related('line_item').get(id=neg_id)
                                
#                                 # Apply values from JSON to the model instance
#                                 # We update BOTH before and after prices as per your JSON structure
#                                 neg_instance.unit_price_before = neg_data.get('unit_price_before')
#                                 neg_instance.unit_price_after = neg_data.get('unit_price_after')
                                
#                                 # If other fields like finalized_comment exist in JSON, apply them too
#                                 if 'finalized_comment' in neg_data:
#                                     neg_instance.finalized_comment = neg_data.get('finalized_comment')

#                                 # CRITICAL: This triggers the model's save() -> calculate_totals()
#                                 # It will use the NEW LineItem quantity and NEW unit prices.
#                                 neg_instance.save()

#                     # --- CASE: TERM ASSIGNMENT ---
#                     elif isinstance(target_obj, TermAssignment):
#                         for field, value in pending_data.items():
#                             if hasattr(target_obj, field):
#                                 setattr(target_obj, field, value)
#                         target_obj.save()

#             return Response({
#                 "message": "Change request approved. Data and calculations synchronized.",
#                 "applied_to": str(instance.content_object)
#             }, status=200)

#         except Exception as e:
#             return Response({"error": f"Synchronization failed: {str(e)}"}, status=400)
class ChangeRequestPatchView(generics.UpdateAPIView):
    queryset = ChangeRequest.objects.all()
    serializer_class = ChangeRequestApprovalSerializer

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        
        if instance.is_approved:
            return Response({"error": "Already approved."}, status=status.HTTP_400_BAD_REQUEST)

        is_approving = request.data.get('is_approved', False)

        try:
            with transaction.atomic():
                # 1. Update the record status (marks it as approved in the CR table)
                serializer = self.get_serializer(instance, data=request.data, partial=True)
                serializer.is_valid(raise_exception=True)
                cr_instance = serializer.save() 

                if is_approving:
                    # 2. Apply the heavy lifting logic from models.py
                    cr_instance.apply_changes()

            return Response({
                "message": "Change request approved and applied successfully."
            }, status=status.HTTP_200_OK)

        except Exception as e:
            # Important: Capture the full error for debugging
            return Response({"error": f"Application failed: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)
import io
import os 
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from xhtml2pdf import pisa

import os
from django.conf import settings

import tempfile
import os
import io
from django.conf import settings
from django.http import HttpResponse
from xhtml2pdf import pisa
import os
import io
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
def link_callback(uri, rel):
    if uri.startswith('/static/'):
        return os.path.join(settings.BASE_DIR, uri.replace('/static/', 'static/'))
    elif uri.startswith('/media/'):
        return os.path.join(settings.BASE_DIR, uri.replace('/media/', 'media/'))
    
    # ✅ ADD THIS (IMPORTANT)
    elif os.path.isabs(uri):
        return uri

    return uri
from weasyprint import HTML
import pathlib

class ProjectBoardPDFExportAPIView(APIView):
    def get(self, request, pb_id):
        pb = get_object_or_404(ProjectBoard.objects.select_related('project', 'buyer'), id=pb_id)
        serializer = PBSummarySerializer(pb)
        data = serializer.data

        # 1. Commercial Requirements Cleaning
        comm_raw = data.get('commercialRequirements') or {}
        clean_comm = {k.replace(' ', '_').replace('/', '_').replace('-', '_'): v for k, v in comm_raw.items()}
        # 2. Comparison Table Flattening (Crucial for the "NoneType" error)
        suppliers = data.get("suppliers") or []
        table_rows = data.get("tableRows") or []
        for row in table_rows:
            aligned = []
            row_map = row.get('suppliers', {})
            for s in suppliers:
                s_id = str(s.get('id'))
                s_entry = row_map.get(s_id, {})
                m = s_entry.get('metrics', {})
                aligned.append({
                    'unit_before': m.get('unitPriceBefore') or 0,
                    'total_before': m.get('totalPriceBefore') or 0,
                    'unit_after': m.get('unitPriceAfter') or 0,
                    'total_after': m.get('totalPriceAfter') or 0
                })
            row['pdf_suppliers'] = aligned


   
        logo_path = os.path.join(settings.BASE_DIR, 'media','t-logo.png')
        logo_uri = pathlib.Path(logo_path).as_uri()
        font_name = 'DejaVu Sans'

    # -----------------------------------
        context = {

            'font_path': os.path.join(settings.BASE_DIR, 'static/fonts/DejaVuSans.ttf'),
            "basicDetails": data.get("basicDetails") or {},
            "commercialRequirements": clean_comm,
            "suppliers": suppliers,
            "tableRows": table_rows,
            "logo_path":logo_uri,
            "customerSelection": data.get("customerSelection") or {},
            "terms": data.get("terms") or {"existing": [], "pending": []},
            "approvers": data.get("approvers") or [],
            "generated_at": timezone.now().strftime("%d-%m-%Y %H:%M"),
            'totals': {
                    'budget': sum(item['budget'] for item in table_rows),
                    'target': sum(item['target'] for item in table_rows),
                    'suppliers': [sum(row['pdf_suppliers'][i]['total_after'] for row in table_rows) for i in range(len(suppliers))]
                }            
        }
        print("======================",context.get("logo_path"),"======================")
        print("======================",context.get("customerSelection"),"======================")

        html_string = render_to_string('project_board.html', context)

        # ✅ CREATE RESPONSE FIRST
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="Project_Board_{pb_id}.pdf"'

        # ✅ GENERATE PDF (NO pisa, NO BytesIO)
        HTML(
            string=html_string,
            base_url=request.build_absolute_uri('/')  # IMPORTANT
        ).write_pdf(response)

        return response
        # result = io.BytesIO()

        # # Generate PDF - Keep the link_callback for your logo/images
        # pisa_status = pisa.pisaDocument(
        #     io.BytesIO(html_string.encode("UTF-8")), 
        #     encoding='UTF-8',
        #     dest=result,
        #     link_callback=link_callback 
        # )

        # if not pisa_status.err:
        #     response = HttpResponse(result.getvalue(), content_type='application/pdf')
        #     response['Content-Disposition'] = f'inline; filename="Project_Board_{pb_id}.pdf"'
        #     return response
        # return HttpResponse("PDF Generation Error", status=400)

from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from rest_framework.views import APIView
import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import io

class ProjectBoardExcelExportAPIView(APIView):
    def get(self, request, pb_id):
        # 1. Get Data
        pb = get_object_or_404(ProjectBoard.objects.select_related('project', 'buyer'), id=pb_id)
        serializer = PBSummarySerializer(pb)
        data = serializer.data

        # 2. Create Workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Procurement Board"
        
        project_details_font_style = Font(name="TKTypeRegular", bold=True, size=18, color="000000")
        
        headers_font_style_bold_14 = Font(name="TKTypeRegular", bold=True, size=14, color="000000")
        headers_font_style_bold_16 = Font(name="TKTypeRegular", bold=True, size=16, color="000000")
        headers_font_style_bold_16_underline = Font(name="TKTypeRegular", bold=True, size=16, color="000000", underline="single")
        headers_font_style_bold_18 = Font(name="TKTypeRegular", bold=True, size=18, color="000000")
        headers_font_style_bold_20 = Font(name="TKTypeRegular", bold=True, size=20, color="000000")
        
        headers_font_style_14 = Font(name="TKTypeRegular", bold=False, size=14, color="000000")
        headers_font_style_16 = Font(name="TKTypeRegular", bold=False, size=16, color="000000")
        headers_font_style_18 = Font(name="TKTypeRegular", bold=False, size=18, color="000000")
        headers_font_style_20 = Font(name="TKTypeRegular", bold=False, size=20, color="000000")
        
        blue_248_248_248 = PatternFill(start_color="f8f8f8", end_color="f8f8f8", fill_type="solid")
        blue_217_222_232 = PatternFill(start_color="d9dee8", end_color="d9dee8", fill_type="solid")
        blue_238_240_242 = PatternFill(start_color="eef0f2", end_color="eef0f2", fill_type="solid")
        blue_242_242_242 = PatternFill(start_color="f2f2f2", end_color="f2f2f2", fill_type="solid")
        
        yellow = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
        
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        
        meduim_border_style = Border(
            left=Side(border_style="medium", color="000000"),
            right=Side(border_style="medium", color="000000"),
            top=Side(border_style="medium", color="000000"),
            bottom=Side(border_style="medium", color="000000")
        )
        
        for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=200):
            for cell in row:
                cell.fill = white_fill
        
        
        column_widths = {
            'A': 12,
            'B': 25,
            'C': 45,
            'D': 59,
            'E': 22,
            'F': 19,
            'G': 18,
            'H': 17,
            'I': 22,
            'J': 26,
            'L': 43,
        }

        row_heights = {
            1: 90.0,
            2: 54.0,
            3: 21.0,
            4: 36.0,
            5: 36.0,
            6: 36.0,
            7: 36.0,
            8: 21.0,
            9: 71.25,
            10: 27.0,
            11: 27.0,
            12: 27.0,
            13: 27.0,
            14: 27.0,
            15: 69.75
        }

        # After you have created your worksheet (ws)
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for row, height in row_heights.items():
            if height is not None:
                ws.row_dimensions[row].height = height
            # None = leave as default (openpyxl will not set it)
        
        # --- PRE-FORMATTING ---
        ws.merge_cells('A1:V1')
        ws['A1'] = "thyssenkrupp Automotive Body Solutions India Pvt. Ltd."
        ws['J2'] = "Procurement Approval Annexure"
        ws['A1'].font = Font(name = "TKTypeRegular",bold=True, size=72, color="00b4f6")
        ws['J2'].font = Font(name="TKTypeRegular", bold=True, size=48, color="305496")
        
        img = Image("tk-logo.png")
        img.anchor = "S1"
        ws.add_image(img)
        
        # --- SECTION 1: BASIC DETAILS ---
        basic = data.get('basicDetails', {})
        ws['C4'] = "*Project Name :"
        ws['C5'] = "*Commodity / Equipment :"
        ws['C6'] = "*Cost Unit/CO Code :"
        ws['K4'] = "*RFQ Sign off :"
        ws['P4'] = "Document No :"
        ws['P5'] = "*Buyers Name :"

        ws['D4'] = basic.get('projectName', "")
        ws['D5'] = basic.get('commodity', "")
        ws['D6'] = basic.get('costUnit', "")
        ws['L4'] = basic.get('rfqSignOff', "")
        ws['Q4'] = basic.get('documentNo', "")
        ws['Q5'] = basic.get('buyerName', "")
        
        # Merge-only cell
        ws.merge_cells('A9:H9')

        sections = [
            ('A10:G10', 'A10', "Currency", headers_font_style_bold_16, "left"),
            ('A11:H11', 'A11', "Commercial Requirements", headers_font_style_bold_16, "left"),
            ('A12:H12', 'A12', "Period of Quotation Validity:", headers_font_style_16, "right"),
            ('A13:H13', 'A13', "Energy Efficiency Requirements:", headers_font_style_16, "right"),
            ('A14:H14', 'A14', " Risk Category/Supplier Qualification:", headers_font_style_16, "right"),
        ]

        for merge_range, cell_ref, value, font_style, align in sections:
            ws.merge_cells(merge_range)
            cell = ws[cell_ref]
            cell.value = value
            cell.font = font_style
            cell.alignment = Alignment(horizontal=align, vertical="center")

        # H10 (Currency value cell)
        cell = ws['H10']
        cell.value = basic.get('currency', "")
        cell.font = headers_font_style_16
        cell.fill = yellow
        cell.alignment = Alignment(vertical="center", horizontal="left")
        

        cells_to_style = ['C4', 'C5', 'C6', 'K4', 'P4', 'P5', 'D4', 'D5', 'D6', 'L4', 'Q4', 'Q5']
        for cell_ref_1 in cells_to_style:
            cell = ws[cell_ref_1]
            cell.font = project_details_font_style
        
        
        commercials = data.get('commercialRequirements', {})
        
        cells_config = [
            ("I9:J10",  "I9",  "Budgets/Targets", 'center'),
            ("I11:J11", "I11", "Expectations", 'center'),
            ("I12:J12", "I12", commercials.get('Period Of Quotation Validity', ""), 'left'),
            ("I13:J13", "I13", commercials.get('Energy Efficiency Requirements', ""), 'left'),
            ("I14:J14", "I14", commercials.get('Risk Category / Supplier Qualification', ""), 'left'),
        ]

        for merge_range, cell_ref, value, align in cells_config:
            ws.merge_cells(merge_range)
            cell = ws[cell_ref]
            cell.value = value
            cell.font = headers_font_style_16
            cell.alignment = Alignment(horizontal=align, vertical='center')
        
        
        headers = [
            ("A15", "Pos"),
            ("B15", "Co no"),
            ("C15", "Item Cd"),
            ("D15", "Description"),
            ("E15", "Unit no"),
            ("F15", "Ref. PO No."),
            ("G15", "Qty."),
            ("H15", "Unit"),
            ("I15", "Budget"),
            ("J15", "Target"),
        ]

        for cell_ref, value in headers:
            cell = ws[cell_ref]
            cell.value = value
            cell.font = headers_font_style_bold_14
            cell.alignment = Alignment(horizontal="center", vertical="center")        
        
        color_blue_248_248_248 = ['A9', 'H9', 'A12', 'A13', 'A14']
        for cell_ref_2 in color_blue_248_248_248:
            cell = ws[cell_ref_2]
            cell.fill = blue_248_248_248
        
        color_blue_217_222_232 = ['A10', 'I9']
        for cell_ref_3 in color_blue_217_222_232:
            cell = ws[cell_ref_3]
            cell.fill = blue_217_222_232
        
        color_blue_238_240_242 = ['A11', 'I11', 'A15','B15','C15','D15','E15','F15','G15','H15','I15','J15',]
        for cell_ref_4 in color_blue_238_240_242:
            cell = ws[cell_ref_4]
            cell.fill = blue_238_240_242


        # --- SECTION 2: SUPPLIER HEADERS (SIDE-BY-SIDE) ---
        suppliers = data.get('suppliers', [])

        BASE_COL = 11   # Column K
        COL_BLOCK = 4   # Each supplier uses 4 columns
        width = 27
        
        # Set width for next 50 columns
        for col_idx in range(BASE_COL, BASE_COL + 50):
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        supplier_col_map = {}

        medium = Side(style="medium", color="000000")
        thin = Side(style="thin", color="000000")
        dotted = Side(style="dotted", color="000000")

        # -- Helper -------------------------------------------------------------
        def apply_border_to_merged_range(ws, start_col, end_col, row, left=None, right=None, top=None, bottom=None):
            """
            Applies border sides correctly across every cell in a merged range.
            - left  : only applied to the first (leftmost) cell
            - right : only applied to the last (rightmost) cell
            - top   : applied to every cell in the range
            - bottom: applied to every cell in the range
            """
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left   = left   if col == start_col else None,
                    right  = right  if col == end_col   else None,
                    top    = top,
                    bottom = bottom
                )
        # -------------------------------------------------------------------------

        for idx, supplier in enumerate(suppliers):
            start_col = BASE_COL + (idx * COL_BLOCK)

            col1 = get_column_letter(start_col)
            col2 = get_column_letter(start_col + 1)
            col3 = get_column_letter(start_col + 2)
            col4 = get_column_letter(start_col + 3)

            # ---------------- Row 9: Supplier Name
            ws.merge_cells(f"{col1}9:{col4}9")
            cell = ws[f"{col1}9"]
            cell.value = supplier.get('supplierName', "")
            cell.font = headers_font_style_bold_20
            cell.fill = blue_217_222_232
            cell.alignment = Alignment(horizontal="center", vertical="top")
            apply_border_to_merged_range(ws, start_col, start_col + 3, 9,
                left=medium, right=medium, top=medium, bottom=medium)

            # ---------------- Row 10: Merged spacer row
            ws.merge_cells(f"{col1}10:{col4}10")
            apply_border_to_merged_range(ws, start_col, start_col + 3, 10,
                left=medium, right=medium, top=medium, bottom=medium)

            # ---------------- Row 11: Before Negotiation
            ws.merge_cells(f"{col1}11:{col2}11")
            cell = ws[f"{col1}11"]
            cell.value = "Before Negotiation"
            cell.font = headers_font_style_16
            cell.fill = blue_238_240_242
            apply_border_to_merged_range(ws, start_col, start_col + 1, 11,
                left=medium, right=medium, top=medium, bottom=dotted)

            # ---------------- Row 11: After Negotiation
            ws.merge_cells(f"{col3}11:{col4}11")
            cell = ws[f"{col3}11"]
            cell.value = "After Negotiation"
            cell.font = headers_font_style_16
            cell.fill = blue_238_240_242
            apply_border_to_merged_range(ws, start_col + 2, start_col + 3, 11,
                left=medium, right=medium, top=medium, bottom=dotted)

            # ---------------- Row 12: Period Of Quotation Validity (Before)
            ws.merge_cells(f"{col1}12:{col2}12")
            cell = ws[f"{col1}12"]
            cell.value = commercials.get('Period Of Quotation Validity', "")
            cell.font = headers_font_style_16
            apply_border_to_merged_range(ws, start_col, start_col + 1, 12,
                left=medium, right=medium, bottom=dotted)

            # ---------------- Row 12: Period Of Quotation Validity (After)
            ws.merge_cells(f"{col3}12:{col4}12")
            cell = ws[f"{col3}12"]
            cell.value = commercials.get('Period Of Quotation Validity', "")
            cell.font = headers_font_style_16
            apply_border_to_merged_range(ws, start_col + 2, start_col + 3, 12,
                left=medium, right=medium, bottom=dotted)

            # ---------------- Row 13: Energy Efficiency Requirements (Before)
            ws.merge_cells(f"{col1}13:{col2}13")
            cell = ws[f"{col1}13"]
            cell.value = commercials.get('Energy Efficiency Requirements', "")
            cell.font = headers_font_style_16
            apply_border_to_merged_range(ws, start_col, start_col + 1, 13,
                left=medium, right=medium, bottom=dotted)

            # ---------------- Row 13: Energy Efficiency Requirements (After)
            ws.merge_cells(f"{col3}13:{col4}13")
            cell = ws[f"{col3}13"]
            cell.value = commercials.get('Energy Efficiency Requirements', "")
            cell.font = headers_font_style_16
            apply_border_to_merged_range(ws, start_col + 2, start_col + 3, 13,
                left=medium, right=medium, bottom=dotted)

            # ---------------- Row 14: Risk Category (Before)
            ws.merge_cells(f"{col1}14:{col2}14")
            cell = ws[f"{col1}14"]
            cell.value = commercials.get('Risk Category / Supplier Qualification', "")
            cell.font = headers_font_style_16
            apply_border_to_merged_range(ws, start_col, start_col + 1, 14,
                left=medium, right=medium, bottom=thin)

            # ---------------- Row 14: Risk Category (After)
            ws.merge_cells(f"{col3}14:{col4}14")
            cell = ws[f"{col3}14"]
            cell.value = commercials.get('Risk Category / Supplier Qualification', "")
            cell.font = headers_font_style_16
            apply_border_to_merged_range(ws, start_col + 2, start_col + 3, 14,
                left=medium, right=medium, bottom=thin)

            # ---------------- Row 15: Final sub-sub headers (single cells, no merge)
            thin_box = Border(left=thin, right=thin, top=thin, bottom=thin)

            cell = ws[f"{col1}15"]
            cell.value = "Unit Price Before Negotiation"
            cell.font = headers_font_style_bold_14
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.fill = blue_238_240_242
            cell.border = thin_box

            cell = ws[f"{col2}15"]
            cell.value = "Total Price Before Negotiation"
            cell.font = headers_font_style_bold_14
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.fill = blue_238_240_242
            cell.border = thin_box

            cell = ws[f"{col3}15"]
            cell.value = "Unit Price After Negotiation"
            cell.font = headers_font_style_bold_14
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.fill = blue_238_240_242
            cell.border = thin_box

            cell = ws[f"{col4}15"]
            cell.value = "Total Price After Negotiation"
            cell.font = headers_font_style_bold_14
            cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
            cell.fill = blue_238_240_242
            cell.border = thin_box
            cell.border = Border(right=Side(border_style="medium", color="000000"))

            # ---------------- Supplier column mapping
            supplier_col_map[str(supplier.get('id'))] = start_col


        # --- SECTION 4: LINE ITEMS (TABLE) ---
        current_row = 16
        table_start_row = current_row  # store starting row

        for item in data.get('tableRows', []):
            budget_value = item.get('budget', '')

            if budget_value:
                budget_value = budget_value[:-7]
                budget_value = int(budget_value)

            # -- Columns 1�10: set values ------------------------------------
            ws.cell(row=current_row, column=1,  value=item.get('pos')).font    = headers_font_style_20
            ws.cell(row=current_row, column=3,  value=item.get('itemCd')).font = headers_font_style_20
            ws.cell(row=current_row, column=4,  value=item.get('desc')).font   = headers_font_style_20
            ws.cell(row=current_row, column=7,  value=item.get('qty')).font    = headers_font_style_20
            ws.cell(row=current_row, column=8,  value=item.get('unit')).font   = headers_font_style_20
            ws.cell(row=current_row, column=9,  value=budget_value).font       = headers_font_style_20
            ws.cell(row=current_row, column=10, value=item.get('target')).font = headers_font_style_20

            # -- Columns 1-10: apply borders ------------------------------------
            for col in range(1, 11):
                cell = ws.cell(row=current_row, column=col)
                cell.border = Border(
                    left   = medium if col == 9  else thin,   # col I  -> medium left
                    right  = medium if col == 10 else thin,   # col J  -> medium right
                    top    = thin,
                    bottom = thin
                )

            # -- Supplier columns: set values + apply borders ---------------------
            item_suppliers = item.get('suppliers', {})

            for s_id, col_idx in supplier_col_map.items():
                metrics = item_suppliers.get(str(s_id), {}).get('metrics', {})
                unit_price_before  = metrics.get('unitPriceBefore',  0)
                total_price_before = metrics.get('totalPriceBefore', 0)
                unit_price_after   = metrics.get('unitPriceAfter',   0)
                total_price_after  = metrics.get('totalPriceAfter',  0)

                ws.cell(row=current_row, column=col_idx,     value=unit_price_before).font  = headers_font_style_20
                ws.cell(row=current_row, column=col_idx + 1, value=total_price_before).font = headers_font_style_20
                ws.cell(row=current_row, column=col_idx + 2, value=unit_price_after).font   = headers_font_style_20
                ws.cell(row=current_row, column=col_idx + 3, value=total_price_after).font  = headers_font_style_20

                # 4 columns per supplier: col_idx, col_idx+1, col_idx+2, col_idx+3
                for offset in range(4):
                    col = col_idx + offset
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = Border(
                        left   = medium if offset == 0 else thin,   # medium on supplier's 1st col
                        right  = medium if offset == 3 else thin,   # medium on supplier's last col
                        top    = thin,
                        bottom = thin
                    )
                    
            current_row += 1

        table_end_row = current_row - 1  # last item row

        current_row = current_row + 3
        print(current_row)

        # -------- DISCOUNT & TOTAL SECTION --------

        discount_rows = [
            ("Discount over all (%)", blue_242_242_242),
            ("Discount over all (net)", blue_242_242_242),
            ("Total", blue_242_242_242),
        ]

        for idx, (label, fill_color) in enumerate(discount_rows):
            row = current_row + idx

            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

            cell = ws.cell(row=row, column=1, value=label)
            cell.font = headers_font_style_bold_16
            cell.fill = fill_color
            cell.alignment = Alignment(vertical="center", horizontal="right")

        # -------- CALCULATE TOTALS FOR SUPPLIERS --------

        total_row = current_row + 2  # "Total" row

        for start_col in supplier_col_map.values():
            for offset in range(4):  # 4 metric columns per supplier
                col = start_col + offset
                column_sum = 0

                for r in range(table_start_row, table_end_row + 1):
                    value = ws.cell(row=r, column=col).value
                    if isinstance(value, (int, float)):
                        column_sum += value

                cell = ws.cell(row=total_row, column=col, value=column_sum)
                cell.font = headers_font_style_16
                cell.border = meduim_border_style

        # -------- CALCULATE TOTAL FOR I & J COLUMNS --------

        total_row = current_row + 2  # "Total" row

        for col in [9, 10]:  # Columns I and J
            column_sum = 0
            has_numeric_value = False

            for r in range(table_start_row, table_end_row + 1):
                value = ws.cell(row=r, column=col).value

                if value is None or value == "":
                    continue
                
                try:
                    numeric_value = float(value)
                    column_sum += numeric_value
                    has_numeric_value = True
                except (ValueError, TypeError):
                    continue  # skip non-numeric strings
                
            if has_numeric_value:
                cell = ws.cell(row=total_row, column=col, value=column_sum)
                cell.font = headers_font_style_16
                cell.border = meduim_border_style
            else:
                ws.cell(row=total_row, column=col, value=None)
        
        ws[f"J{current_row + 2}"].border = meduim_border_style

        # -------- COMMERCIAL SECTION --------

        commercial_rows = [
            ("Customer Pre-selected:", "Customer Pre-Selected"),
            ("Restrictions Functional Department:", "Restrictions Functional Department"),
            ("Pre-negotiated:", "Pre-negotiated"),
            ("According To Customer Preferred List:", "According Customer Preferred List"),
        ]

        for idx, (label, key) in enumerate(commercial_rows):
            row = current_row + 3 + idx

            # Left merged label (A-H)
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)

            cell = ws.cell(row=row, column=1, value=label)
            cell.font = headers_font_style_bold_16
            cell.fill = blue_238_240_242
            cell.alignment = Alignment(vertical="center", horizontal="right")

            # Value merged (I-J)
            ws.merge_cells(start_row=row, start_column=9, end_row=row, end_column=10)

            cell = ws.cell(row=row, column=9, value=commercials.get(key, ""))
            cell.font = headers_font_style_16
            cell.fill = blue_238_240_242
            cell.alignment = Alignment(vertical="center", horizontal="left")


        # -------- SUPPLIER BLOCK MERGE + BORDER --------

        for start_col in supplier_col_map.values():
            for offset in range(3, 7):
                row = current_row + offset

                # Apply border to full width
                for col in range(start_col, start_col + 4):
                    ws.cell(row=row, column=col).border = meduim_border_style

                # Merge supplier block
                ws.merge_cells(
                    start_row=row,
                    start_column=start_col,
                    end_row=row,
                    end_column=start_col + 3
                )

                ws.cell(row=row, column=start_col).fill = blue_238_240_242
        
        ws[f"A{current_row + 8}"] = "Terms & Conditions :"
        ws[f"A{current_row + 8}"].font = headers_font_style_bold_16_underline
        
        base_row = current_row + 9
        
        left_terms = [
            "Buy Off Criteria",
            "Delivery",
            "Warranty",
            "Incoterms",
            "Documentation",
            "Payment Terms",
            "Others if Any",
            "References",
        ]
        
        right_terms = [
            "Old Ref. / Offer",
            "*HS / Teriff Code :",
            "Local Taxes",
            "Quality Standards",
            "Warranty for Defects",
            "EU/USA Sanctions",
            "Spare part list",
            "Service Support",
        ]

        terms_data = {
            t["term"]: t
            for t in data.get("terms", {}).get("existing", [])
        }
        
        # ---- LEFT BLOCK (B, C, D) ----
        row_pointer = base_row

        for term in left_terms:
            term_data = terms_data.get(term, {})
            expected = term_data.get("expected", {})
            agreed = term_data.get("agreed", {})

            # Special logic for Delivery & Payment Terms
            if term in ["Delivery", "Payment Terms"]:
                expected_value = ""
                agreed_value = ""

                if expected.get("percentage") is not None and expected.get("date"):
                    exp_date = datetime.strptime(expected["date"], "%Y-%m-%d")
                    formatted_exp_date = exp_date.strftime("%d %b %Y")
                    expected_value = f"{expected['percentage']}% at {formatted_exp_date}"

                if agreed.get("percentage") is not None and agreed.get("date"):
                    agr_date = datetime.strptime(agreed["date"], "%Y-%m-%d")
                    formatted_agr_date = agr_date.strftime("%d %b %Y")
                    agreed_value = f"{agreed['percentage']}% at {formatted_agr_date}"

            else:
                expected_value = expected.get("text_detail", "")
                agreed_value = agreed.get("text_detail", "")

            # Term in Column B
            cell = ws.cell(row=row_pointer, column=2, value=term)
            cell.font = headers_font_style_bold_16
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Expected label (Column C)
            cell = ws.cell(row=row_pointer, column=3, value="Expected: ")
            cell.font = headers_font_style_16
            cell.alignment = Alignment(horizontal="right", vertical="center")

            # Expected value (Column D)
            ws.merge_cells(start_row=row_pointer, start_column=4, end_row=row_pointer, end_column=8)
            cell = ws.cell(row=row_pointer, column=4, value=expected_value)
            cell.font = headers_font_style_14
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Agreed label (Column C next row)
            cell = ws.cell(row=row_pointer + 1, column=3, value="Agreed: ")
            cell.font = headers_font_style_bold_16
            cell.alignment = Alignment(horizontal="right", vertical="center")

            # Agreed value (Column D next row)
            ws.merge_cells(start_row=row_pointer + 1, start_column=4, end_row=row_pointer + 1, end_column=8)
            cell = ws.cell(row=row_pointer + 1, column=4, value=agreed_value)
            cell.font = headers_font_style_14
            cell.alignment = Alignment(horizontal="left", vertical="center")

            row_pointer += 3

        # ---- RIGHT BLOCK (L, M, N) ----
        row_pointer = base_row
        
        for term in right_terms:
            term_data = terms_data.get(term, {})
            expected = term_data.get("expected", {})
            agreed = term_data.get("agreed", {})
        
            # No special case needed here (none are Delivery/Payment Terms)
            expected_value = expected.get("text_detail", "")
            agreed_value = agreed.get("text_detail", "")
        
            # Term in Column L
            cell = ws.cell(row=row_pointer, column=12, value=term)
            cell.font = headers_font_style_bold_16
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
            # Expected label (Column M)
            cell = ws.cell(row=row_pointer, column=13, value="Expected: ")
            cell.font = headers_font_style_16
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
            # Expected value (Column N)
            ws.merge_cells(start_row=row_pointer, start_column=14, end_row=row_pointer, end_column=18)
            cell = ws.cell(row=row_pointer, column=14, value=expected_value)
            cell.font = headers_font_style_14
            cell.alignment = Alignment(horizontal="left", vertical="center")
        
            # Agreed label (Column M next row)
            cell = ws.cell(row=row_pointer + 1, column=13, value="Agreed: ")
            cell.font = headers_font_style_bold_16
            cell.alignment = Alignment(horizontal="right", vertical="center")
        
            # Agreed value (Column N next row)
            ws.merge_cells(start_row=row_pointer + 1, start_column=14, end_row=row_pointer + 1, end_column=18)
            cell = ws.cell(row=row_pointer + 1, column=14, value=agreed_value)
            cell.font = headers_font_style_14
            cell.alignment = Alignment(horizontal="left", vertical="center")

            row_pointer += 3

            approvers = data.get("approvers", [])

            project_manager_name = None
            md_name = None
            hod_name = None
            superuser_name = None
            employee_name = None
            admin_name = None
            deputy_manager_name = None
            assistant_manager_name = None

            for approver in approvers:
                role = approver.get("role")

                if role == "project_manager":
                    project_manager_name = approver.get("approver_name")

                elif role == "md":
                    md_name = approver.get("approver_name")
                
                elif role == "superuser":
                    superuser_name = approver.get("superuser")
                
                elif role == "admin":
                    admin_name = approver.get("admin")
                
                elif role == "employee":
                    employee_name = approver.get("employee")
                
                elif role == "hod":
                    hod_name = approver.get("hod")
                
                elif role == "deputy_manager":
                    deputy_manager_name = approver.get("deputy_manager")
                
                elif role == "assistant_manager":
                    assistant_manager_name = approver.get("sassistant_managerr")
                

        base_row = base_row + 30
        
        ws[f"A{base_row}"] = "Presented By"
        ws[f"A{base_row}"].font = headers_font_style_bold_14
        ws[f"A{base_row}"].alignment = Alignment(vertical="center", horizontal="left")

        ws[f"C{base_row}"] = "Project Cost Controlling"
        ws[f"C{base_row}"].font = headers_font_style_bold_14
        ws[f"C{base_row}"].alignment = Alignment(vertical="center", horizontal="left")

        ws[f"D{base_row}"] = "User Dept. Rep."
        ws[f"D{base_row}"].font = headers_font_style_bold_14
        ws[f"D{base_row}"].alignment = Alignment(vertical="center", horizontal="left")

        # Project Manager Name

        ws[f"E{base_row - 1}"].value = project_manager_name
        ws[f"E{base_row - 1}"].font = headers_font_style_14
        ws[f"E{base_row - 1}"].alignment = Alignment(vertical="center", horizontal="center")
        
        ws[f"E{base_row}"] = "Project Manager"
        ws[f"E{base_row}"].font = headers_font_style_bold_14
        ws[f"E{base_row}"].alignment = Alignment(vertical="center", horizontal="left")

        # HOD Name

        ws[f"G{base_row - 1}"].value = hod_name
        ws[f"G{base_row - 1}"].font = headers_font_style_14
        ws[f"G{base_row - 1}"].alignment = Alignment(vertical="center", horizontal="center")

        ws[f"G{base_row}"] = "Head of Dept."
        ws[f"G{base_row}"].font = headers_font_style_bold_14
        ws[f"G{base_row}"].alignment = Alignment(vertical="center", horizontal="center")

        ws[f"I{base_row}"] = "Head of Sales & CE"
        ws[f"I{base_row}"].font = headers_font_style_bold_14
        ws[f"I{base_row}"].alignment = Alignment(vertical="center", horizontal="center")

        ws[f"L{base_row}"] = "Head Engineering and Projects"
        ws[f"L{base_row}"].font = headers_font_style_bold_14
        ws[f"L{base_row}"].alignment = Alignment(vertical="center", horizontal="center")

        ws[f"O{base_row}"] = "Head PSM & Manufacturing"
        ws[f"O{base_row}"].font = headers_font_style_bold_14
        ws[f"O{base_row}"].alignment = Alignment(vertical="center", horizontal="center")

        ws[f"R{base_row}"] = "Head Finance & Controlling "
        ws[f"R{base_row}"].font = headers_font_style_bold_14
        ws[f"R{base_row}"].alignment = Alignment(vertical="center", horizontal="center")

        # MD Name

        ws[f"U{base_row - 1}"].value = md_name
        ws[f"U{base_row - 1}"].font = headers_font_style_14
        ws[f"U{base_row - 1}"].alignment = Alignment(vertical="center", horizontal="center")
        
        ws[f"U{base_row}"] = "Managing Director & CEO"
        ws[f"U{base_row}"].font = headers_font_style_bold_14
        ws[f"U{base_row}"].alignment = Alignment(vertical="center", horizontal="center")
        
        
        ws[f"A{base_row + 2}"] = "Important Guidelines :"
        ws[f"A{base_row + 2}"].font = headers_font_style_18
        
        ws[f"A{base_row + 3}"] = '1) All Fields marked with " * " are mandatory to be filled with correct and latest discussions and agreement with suppliers.'
        ws[f"A{base_row + 3}"].font = headers_font_style_16

        ws[f"A{base_row + 4}"] = "2) Cash flow / Revenue sheet to be updated before raising requiest for Procurement / Sourcing Board."
        ws[f"A{base_row + 4}"].font = headers_font_style_16

        ws[f"A{base_row + 5}"] = "3) Purchase Order / LOI / Final confirmation to be shared with supplier only after approval from Procurement / Sourcing Boards."
        ws[f"A{base_row + 5}"].font = headers_font_style_16

        ws[f"K{base_row + 3}"] = "4) Back-up to be provided for all agreements."
        ws[f"K{base_row + 3}"].font = headers_font_style_16

        ws[f"K{base_row + 4}"] = "5) Agreements are expected to be aligned with Project requirements / expectatations."
        ws[f"K{base_row + 4}"].font = headers_font_style_16

        
        top_medium = Side(border_style="medium", color="000000")
        right_medium = Side(border_style="medium", color="000000")
        bottom_thin = Side(border_style="thin", color="000000")
        
        # Top border (apply to all cells A9 ? H9)
        for col in range(1, 9):  # A=1, H=8
            ws.cell(row=9, column=col).border = Border(top=top_medium)
        
        # Bottom border (apply to all cells A9 ? H9)
        for col in range(1, 9):
            existing = ws.cell(row=9, column=col).border
            ws.cell(row=9, column=col).border = Border(
                top=existing.top,
                bottom=bottom_thin
            )
        
        # Right border (apply ONLY to H9)
        existing = ws.cell(row=9, column=8).border
        ws.cell(row=9, column=8).border = Border(
            top=existing.top,
            bottom=existing.bottom,
            right=right_medium
        )

        # Top border (I9:J9)
        for col in range(9, 11):
            ws.cell(row=9, column=col).border = Border(top=medium)

        # Bottom border (I10:J10)
        for col in range(9, 11):
            existing = ws.cell(row=10, column=col).border
            ws.cell(row=10, column=col).border = Border(
                top=existing.top,
                bottom=medium
            )

        # Left border (I9:I10)
        for row in range(9, 11):
            existing = ws.cell(row=row, column=9).border
            ws.cell(row=row, column=9).border = Border(
                top=existing.top,
                bottom=existing.bottom,
                left=medium
            )

        # Right border (J9:J10)
        for row in range(9, 11):
            existing = ws.cell(row=row, column=10).border
            ws.cell(row=row, column=10).border = Border(
                top=existing.top,
                bottom=existing.bottom,
                left=existing.left,
                right=medium
            )

        # Bottom border for merged A10:G10
        for col in range(1, 8):  # A=1 to G=7
            existing = ws.cell(row=10, column=col).border
            ws.cell(row=10, column=col).border = Border(
                top=existing.top,
                left=existing.left,
                right=existing.right,
                bottom=thin
            )

        # Bottom border for H10
        existing = ws.cell(row=10, column=8).border
        ws.cell(row=10, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=existing.right,
            bottom=thin
        )

        # Rows 11, 12, 13 -> medium left/right, dotted bottom
        for row in [11, 12, 13]:
            # Left border (Column I)
            existing = ws.cell(row=row, column=9).border
            ws.cell(row=row, column=9).border = Border(
                top=existing.top,
                left=medium,
                right=existing.right,
                bottom=dotted
            )

            # Right border (Column J)
            existing = ws.cell(row=row, column=10).border
            ws.cell(row=row, column=10).border = Border(
                top=existing.top,
                left=existing.left,
                right=medium,
                bottom=dotted
            )

        # Row 14 -> medium left/right/bottom
        row = 14

        # Left border (Column I)
        existing = ws.cell(row=row, column=9).border
        ws.cell(row=row, column=9).border = Border(
            top=existing.top,
            left=medium,
            right=existing.right,
            bottom=thin
        )

        # Right border (Column J)
        existing = ws.cell(row=row, column=10).border
        ws.cell(row=row, column=10).border = Border(
            top=existing.top,
            left=existing.left,
            right=medium,
            bottom=thin
        )

        # A15 to G15
        for col in range(1, 8):
            ws.cell(row=15, column=col).border = Border(
                top=thin,
                right=thin,
                bottom=thin
            )
        
        # H15
        ws.cell(row=15, column=8).border = Border(
            top=thin,
            right=medium,
            bottom=thin
        )
        
        # I15
        ws.cell(row=15, column=9).border = Border(
            top=thin,
            right=thin,
            bottom=thin
        )
        
        # J15
        ws.cell(row=15, column=10).border = Border(
            top=thin,
            right=medium,
            bottom=thin
        )

        row = current_row

        # Top border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=medium,
                left=existing.left,
                right=existing.right,
                bottom=existing.bottom
            )

        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )

        row = current_row + 1

        # Bottom border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=existing.top,
                left=existing.left,
                right=existing.right,
                bottom=medium
            )

        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )
        
        row = current_row + 2
        
        # Bottom border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=medium,
                left=existing.left,
                right=existing.right,
                bottom=medium
            )
        
        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )

        row = current_row + 3
        # Bottom border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=existing.top,
                left=existing.left,
                right=existing.right,
                bottom=dotted
            )
        
        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )


        row = current_row + 4
        # Bottom border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=existing.top,
                left=existing.left,
                right=existing.right,
                bottom=dotted
            )
        
        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )

        row = current_row + 5
        # Bottom border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=existing.top,
                left=existing.left,
                right=existing.right,
                bottom=dotted
            )
        
        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )

        row = current_row + 6
        # Bottom border across A -> H
        for col in range(1, 9):
            existing = ws.cell(row=row, column=col).border
            ws.cell(row=row, column=col).border = Border(
                top=existing.top,
                left=existing.left,
                right=existing.right,
                bottom=medium
            )
        
        # Right border on H only
        existing = ws.cell(row=row, column=8).border
        ws.cell(row=row, column=8).border = Border(
            top=existing.top,
            left=existing.left,
            right=thin,
            bottom=existing.bottom
        )

        rows_config = [
            (current_row + 3, medium, dotted),  # top, bottom
            (current_row + 4, dotted, dotted),
            (current_row + 5, dotted, dotted),
            (current_row + 6, dotted, medium),
        ]

        for row, top_side, bottom_side in rows_config:
            # Left cell (Column I = 9)
            ws.cell(row=row, column=9).border = Border(
                top=top_side,
                left=thin,
                bottom=bottom_side
            )

            # Right cell (Column J = 10)
            ws.cell(row=row, column=10).border = Border(
                top=top_side,
                right=medium,
                bottom=bottom_side
            )

        # # 3. Create File Response
        # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        # response['Content-Disposition'] = f'attachment; filename=Procurement_Board_{pb_id}.xlsx'
        # wb.save(response)
        # return response

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        # 4. Fetch All ProcurementUpload Files
        attachments = ProcurementUpload.objects.filter(pb=pb)

        # 5. Create ZIP in Memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            # Add Excel File
            excel_filename = f"Procurement_Board_{pb_id}.xlsx"
            zf.writestr(excel_filename, excel_buffer.getvalue())

            # Add Attachments
            for upload in attachments:
                if upload.attachment:
                    try:
                        # Extract original filename
                        original_name = upload.attachment.name.split('/')[-1]
                        # Read the file content from storage
                        zf.writestr(f"attachments/{original_name}", upload.attachment.read())
                    except Exception as e:
                        print(f"Error zipping file: {e}")

        # 6. Return ZIP Response
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename=Project_Board_Package_{pb_id}.zip'
        return response


from django.views import View

class ProjectProcurementPlanListView(View):
    def get(self, request, project_id):
        # 1. Fetch all 'commodity_equipment' names from the ProjectBoard (the Master Plan)
        # We get a flat list of strings to make searching fast
        planned_item_names = ProjectBoard.objects.filter(
            project_id=project_id
        ).values_list('commodity_equipment', flat=True)

        # 2. Fetch all records from ExcelFile for this project
        excel_records = ExcelFile.objects.filter(project_id=project_id)

        # 3. Build the results based on the Excel records
        results = []
        for record in excel_records:
            # Check if this specific Excel item name exists in the ProjectBoard plan
            is_planned = record.item in planned_item_names
            status = "procured" if is_planned else "not in plan"
            
            results.append({
                "id": record.id,
                "item_name": record.item,
                "item_budget": record.item_budget,
                "total_cost": record.total,
                "delivery_date": record.delivery_date,
                "status": status,
                "project_id": project_id
            })

        # 4. Return as JSON
        return JsonResponse(results, safe=False)


import requests

class CurrencyExchangeView(APIView):
    def get(self, request):
        # Your specific API Key and URL (Base is INR)
        api_url = "https://v6.exchangerate-api.com/v6/40c3bb8d90ad164ae82f4714/latest/INR"
        
        try:
            # response = requests.get(api_url, timeout=10)
            # data = response.json()

            # if data.get('result') == 'success':
            
            # Logic: Since the API base is INR, USD is a decimal (e.g., 0.012).
            # To get "1 USD = X", we divide the target rate by the USD rate.

            # Calculated conversions
            payload = {
                "base_currency": "INR",
                "rates_in_inr": {
                    # 1 USD to INR (e.g., ~83.00)
                    "INR": 1,

                    "USD": 92,
                    # 1 EUR to INR (e.g., ~90.00)
                    "EUR": 106,
                    # 1 CNY to INR (e.g., ~11.00)
                    "CNY": 13,
                    # 1 JPY to INR (e.g., ~0.55)
                    "JPY": 0.58,
                },
            }
            return Response(payload, status=status.HTTP_200_OK)
            

        except requests.exceptions.RequestException as e:
            return Response({"error": "Connection failed", "details": str(e)}, status=status.HTTP_503_SERVICE_UNAVAILABLE)                
